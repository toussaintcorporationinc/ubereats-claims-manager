import csv
import io
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from fastapi import UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.auth import can_access_restaurant
from app.core.config import get_settings
from app.models import ClaimOrder, ImportBatch, ImportRow, Restaurant, User
from app.models.domain import utc_now
from app.services.audit import add_audit_log

SUPPORTED_IMPORT_EXTENSIONS = {".csv": "csv", ".xlsx": "xlsx"}
PREVIEW_ROW_LIMIT = 50

HEADER_ALIASES = {
    "restaurant_id": "restaurant_id",
    "id_restaurant": "restaurant_id",
    "restaurant_name": "restaurant_name",
    "nom_restaurant": "restaurant_name",
    "uber_order_number": "uber_order_number",
    "order_number": "uber_order_number",
    "numero_commande": "uber_order_number",
    "commande_uber": "uber_order_number",
    "uber_id": "uber_order_number",
    "customer_name": "customer_name",
    "client": "customer_name",
    "nom_client": "customer_name",
    "order_date": "order_date",
    "date": "order_date",
    "date_commande": "order_date",
    "order_time": "order_time",
    "heure": "order_time",
    "heure_commande": "order_time",
    "cancellation_time": "cancellation_time",
    "heure_annulation": "cancellation_time",
    "annulation": "cancellation_time",
    "order_amount": "order_amount",
    "amount": "order_amount",
    "montant": "order_amount",
    "montant_commande": "order_amount",
    "total": "order_amount",
    "currency": "currency",
    "devise": "currency",
    "accepted_by_restaurant": "accepted_by_restaurant",
    "acceptee": "accepted_by_restaurant",
    "commande_acceptee": "accepted_by_restaurant",
    "prepared_before_cancellation": "prepared_before_cancellation",
    "preparee": "prepared_before_cancellation",
    "commande_preparee": "prepared_before_cancellation",
    "prete": "prepared_before_cancellation",
    "loss_type": "loss_type",
    "type_perte": "loss_type",
    "perte": "loss_type",
    "gaspillage": "loss_type",
    "notes": "notes",
    "commentaire": "notes",
    "commentaires": "notes",
    "internal_reference": "internal_reference",
    "reference_interne": "internal_reference",
}

TEXT_FIELDS = ("internal_reference", "uber_order_number", "customer_name", "loss_type", "notes")
FINAL_BATCH_STATUSES = {"confirmed", "partially_imported", "failed"}


class OrderImportError(Exception):
    def __init__(self, message: str, status_code: int = status.HTTP_400_BAD_REQUEST) -> None:
        self.message = message
        self.status_code = status_code
        super().__init__(message)


@dataclass(frozen=True)
class UploadedImportFile:
    original_filename: str
    file_type: str
    content: bytes


@dataclass(frozen=True)
class RawImportRow:
    row_number: int
    raw_data: dict[str, str]


def ensure_import_storage() -> Path:
    storage_dir = get_settings().import_storage_dir
    storage_dir.mkdir(parents=True, exist_ok=True)
    return storage_dir


def create_order_import_preview(db: Session, user: User, upload_file: UploadFile) -> ImportBatch:
    uploaded_file = read_import_upload(upload_file)
    raw_rows = parse_import_rows(uploaded_file)
    batch = ImportBatch(
        uploaded_by_user_id=user.id,
        original_filename=uploaded_file.original_filename,
        file_type=uploaded_file.file_type,
        status="uploaded",
    )
    db.add(batch)
    db.flush()

    seen_order_keys: set[tuple[int, str]] = set()
    import_rows = [
        build_import_row(db, user, batch.id, raw_row, seen_order_keys)
        for raw_row in raw_rows
    ]
    db.add_all(import_rows)
    db.flush()

    apply_batch_counts(batch, import_rows)
    batch.status = "parsed"
    add_audit_log(
        db,
        entity_type="import_batch",
        entity_id=batch.id,
        action="import_batch.previewed",
        user_id=user.id,
        new_value={
            "original_filename": batch.original_filename,
            "file_type": batch.file_type,
            "total_rows": batch.total_rows,
            "valid_rows": batch.valid_rows,
            "invalid_rows": batch.invalid_rows,
            "duplicate_rows": batch.duplicate_rows,
            "unauthorized_rows": batch.unauthorized_rows,
        },
    )
    return batch


def read_import_upload(upload_file: UploadFile) -> UploadedImportFile:
    original_filename = safe_original_filename(upload_file.filename)
    extension = Path(original_filename).suffix.lower()
    file_type = SUPPORTED_IMPORT_EXTENSIONS.get(extension)
    if file_type is None:
        raise OrderImportError("Import file extension is not allowed")

    max_size = get_settings().import_max_file_size_mb * 1024 * 1024
    content = upload_file.file.read(max_size + 1)
    if not content:
        raise OrderImportError("Import file cannot be empty")
    if len(content) > max_size:
        raise OrderImportError("Import file is too large", status.HTTP_413_CONTENT_TOO_LARGE)

    ensure_import_storage()
    return UploadedImportFile(original_filename=original_filename, file_type=file_type, content=content)


def parse_import_rows(uploaded_file: UploadedImportFile) -> list[RawImportRow]:
    if uploaded_file.file_type == "csv":
        return parse_csv_rows(uploaded_file.content)
    return parse_xlsx_rows(uploaded_file.content)


def parse_csv_rows(content: bytes) -> list[RawImportRow]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise OrderImportError("Import file must include a header row")

    rows: list[RawImportRow] = []
    for row_number, row in enumerate(reader, start=2):
        raw_data = {
            str(header).strip(): stringify_cell(value)
            for header, value in row.items()
            if header is not None and str(header).strip()
        }
        if not is_empty_raw_row(raw_data):
            rows.append(RawImportRow(row_number=row_number, raw_data=raw_data))
    return rows


def parse_xlsx_rows(content: bytes) -> list[RawImportRow]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    row_iterator = sheet.iter_rows(values_only=True)
    try:
        headers = next(row_iterator)
    except StopIteration as exc:
        raise OrderImportError("Import file must include a header row") from exc

    header_values = [stringify_cell(header) for header in headers]
    if not any(header_values):
        raise OrderImportError("Import file must include a header row")

    rows: list[RawImportRow] = []
    for row_number, values in enumerate(row_iterator, start=2):
        raw_data = {
            header: stringify_cell(value)
            for header, value in zip(header_values, values)
            if header
        }
        if not is_empty_raw_row(raw_data):
            rows.append(RawImportRow(row_number=row_number, raw_data=raw_data))
    return rows


def build_import_row(
    db: Session,
    user: User,
    batch_id: int,
    raw_row: RawImportRow,
    seen_order_keys: set[tuple[int, str]],
) -> ImportRow:
    normalized_data, row_status, errors, warnings = normalize_import_row(db, user, raw_row.raw_data, seen_order_keys)
    return ImportRow(
        batch_id=batch_id,
        row_number=raw_row.row_number,
        raw_data=raw_row.raw_data,
        normalized_data=normalized_data,
        status=row_status,
        errors=errors,
        warnings=warnings,
    )


def normalize_import_row(
    db: Session,
    user: User,
    raw_data: dict[str, str],
    seen_order_keys: set[tuple[int, str]],
) -> tuple[dict[str, Any] | None, str, list[str], list[str]]:
    canonical_data = canonicalize_raw_data(raw_data)
    errors: list[str] = []
    warnings: list[str] = []
    normalized: dict[str, Any] = {}

    restaurant = resolve_restaurant(db, canonical_data, errors)
    if restaurant is not None:
        normalized["restaurant_id"] = restaurant.id
        if not can_access_restaurant(db, user, restaurant.id):
            errors.append("restaurant_access_denied")
            return normalized, "unauthorized", errors, warnings

    uber_order_number = clean_optional(canonical_data.get("uber_order_number"))
    if not uber_order_number:
        errors.append("missing_uber_order_number")
    else:
        normalized["uber_order_number"] = uber_order_number

    amount_text = clean_optional(canonical_data.get("order_amount"))
    if not amount_text:
        errors.append("missing_order_amount")
    else:
        amount = parse_amount(amount_text)
        if amount is None:
            errors.append("invalid_order_amount")
        else:
            normalized["order_amount"] = str(amount)

    currency = clean_optional(canonical_data.get("currency"))
    if currency:
        currency = currency.upper()
        if len(currency) != 3:
            errors.append("invalid_currency")
        else:
            normalized["currency"] = currency
    else:
        normalized["currency"] = "EUR"
        warnings.append("currency_defaulted_to_EUR")

    normalize_optional_date(canonical_data, normalized, errors, "order_date")
    normalize_optional_time(canonical_data, normalized, errors, "order_time")
    normalize_optional_time(canonical_data, normalized, errors, "cancellation_time")
    normalize_optional_bool(canonical_data, normalized, warnings, "accepted_by_restaurant", default=True)
    normalize_optional_bool(canonical_data, normalized, warnings, "prepared_before_cancellation", default=True)

    for field in TEXT_FIELDS:
        if value := clean_optional(canonical_data.get(field)):
            normalized[field] = value

    if errors:
        return normalized or None, "invalid", errors, warnings

    restaurant_id = normalized["restaurant_id"]
    order_number = normalized["uber_order_number"]
    order_key = (restaurant_id, order_number)
    if order_exists(db, restaurant_id, order_number):
        return normalized, "duplicate", ["duplicate_existing_order"], warnings
    if order_key in seen_order_keys:
        return normalized, "duplicate", ["duplicate_in_file"], warnings
    seen_order_keys.add(order_key)

    return normalized, "valid", [], warnings


def canonicalize_raw_data(raw_data: dict[str, str]) -> dict[str, str]:
    canonical_data: dict[str, str] = {}
    ambiguous_restaurant: str | None = None
    for header, value in raw_data.items():
        normalized_header = normalize_header(header)
        if normalized_header == "restaurant":
            ambiguous_restaurant = clean_optional(value)
            continue
        canonical_name = HEADER_ALIASES.get(normalized_header)
        if canonical_name and canonical_name not in canonical_data:
            canonical_data[canonical_name] = value.strip()

    if ambiguous_restaurant and "restaurant_id" not in canonical_data and "restaurant_name" not in canonical_data:
        if ambiguous_restaurant.isdigit():
            canonical_data["restaurant_id"] = ambiguous_restaurant
        else:
            canonical_data["restaurant_name"] = ambiguous_restaurant
    return canonical_data


def resolve_restaurant(db: Session, canonical_data: dict[str, str], errors: list[str]) -> Restaurant | None:
    restaurant_id_text = clean_optional(canonical_data.get("restaurant_id"))
    restaurant_name = clean_optional(canonical_data.get("restaurant_name"))
    if not restaurant_id_text and not restaurant_name:
        errors.append("missing_restaurant")
        return None

    if restaurant_id_text:
        try:
            restaurant_id = int(restaurant_id_text)
        except ValueError:
            errors.append("invalid_restaurant_id")
            return None
        restaurant = db.get(Restaurant, restaurant_id)
    else:
        normalized_name = restaurant_name.casefold() if restaurant_name else ""
        restaurant = db.scalar(
            select(Restaurant)
            .where(func.lower(Restaurant.name) == normalized_name.lower())
            .order_by(Restaurant.id)
        )

    if restaurant is None:
        errors.append("restaurant_not_found")
    return restaurant


def confirm_order_import_batch(db: Session, batch: ImportBatch, user: User) -> tuple[int, int, list[str]]:
    if batch.status in FINAL_BATCH_STATUSES:
        raise OrderImportError("Import batch has already been confirmed", status.HTTP_409_CONFLICT)
    if batch.status == "cancelled":
        raise OrderImportError("Import batch is cancelled", status.HTTP_409_CONFLICT)

    rows = list(db.scalars(select(ImportRow).where(ImportRow.batch_id == batch.id).order_by(ImportRow.id)).all())
    valid_rows = [row for row in rows if row.status == "valid"]
    errors: list[str] = []
    created_orders_count = 0

    for row in valid_rows:
        normalized_data = row.normalized_data or {}
        restaurant_id = int(normalized_data["restaurant_id"])
        uber_order_number = str(normalized_data["uber_order_number"])
        if not can_access_restaurant(db, user, restaurant_id):
            mark_row_skipped(row, "restaurant_access_denied_at_confirm")
            errors.append(f"row {row.row_number}: restaurant_access_denied_at_confirm")
            continue
        if order_exists(db, restaurant_id, uber_order_number):
            mark_row_skipped(row, "duplicate_existing_order_at_confirm")
            errors.append(f"row {row.row_number}: duplicate_existing_order_at_confirm")
            continue

        order = ClaimOrder(
            restaurant_id=restaurant_id,
            internal_reference=normalized_data.get("internal_reference"),
            uber_order_number=uber_order_number,
            customer_name=normalized_data.get("customer_name"),
            order_date=parse_confirm_date(normalized_data.get("order_date")),
            order_time=parse_confirm_time(normalized_data.get("order_time")),
            cancellation_time=parse_confirm_time(normalized_data.get("cancellation_time")),
            order_amount=Decimal(str(normalized_data["order_amount"])),
            currency=str(normalized_data.get("currency") or "EUR"),
            accepted_by_restaurant=normalized_data.get("accepted_by_restaurant"),
            prepared_before_cancellation=normalized_data.get("prepared_before_cancellation"),
            loss_type=normalized_data.get("loss_type"),
            notes=normalized_data.get("notes"),
        )
        db.add(order)
        db.flush()
        row.status = "created"
        row.created_order_id = order.id
        created_orders_count += 1
        add_audit_log(
            db,
            entity_type="claim_order",
            entity_id=order.id,
            action="claim_order.created_from_import",
            user_id=user.id,
            new_value={
                "import_batch_id": batch.id,
                "import_row_id": row.id,
                "restaurant_id": order.restaurant_id,
                "uber_order_number": order.uber_order_number,
                "order_amount": order.order_amount,
                "currency": order.currency,
                "status": order.status,
            },
        )

    batch.created_orders_count = created_orders_count
    batch.confirmed_at = utc_now()
    if created_orders_count == len(valid_rows) and not errors:
        batch.status = "confirmed"
    elif created_orders_count > 0:
        batch.status = "partially_imported"
    else:
        batch.status = "failed"
    batch.updated_at = utc_now()

    skipped_rows = batch.total_rows - created_orders_count
    add_audit_log(
        db,
        entity_type="import_batch",
        entity_id=batch.id,
        action="import_batch.confirmed",
        user_id=user.id,
        new_value={
            "status": batch.status,
            "created_orders_count": created_orders_count,
            "skipped_rows": skipped_rows,
            "errors": errors,
        },
    )
    return created_orders_count, skipped_rows, errors


def cancel_order_import_batch(db: Session, batch: ImportBatch, user: User) -> None:
    if batch.status in FINAL_BATCH_STATUSES:
        raise OrderImportError("Confirmed import batches cannot be cancelled", status.HTTP_409_CONFLICT)
    if batch.status == "cancelled":
        raise OrderImportError("Import batch is already cancelled", status.HTTP_409_CONFLICT)
    batch.status = "cancelled"
    batch.updated_at = utc_now()
    add_audit_log(
        db,
        entity_type="import_batch",
        entity_id=batch.id,
        action="import_batch.cancelled",
        user_id=user.id,
        new_value={"status": batch.status},
    )


def apply_batch_counts(batch: ImportBatch, rows: list[ImportRow]) -> None:
    batch.total_rows = len(rows)
    batch.valid_rows = sum(1 for row in rows if row.status == "valid")
    batch.invalid_rows = sum(1 for row in rows if row.status == "invalid")
    batch.duplicate_rows = sum(1 for row in rows if row.status == "duplicate")
    batch.unauthorized_rows = sum(1 for row in rows if row.status == "unauthorized")


def mark_row_skipped(row: ImportRow, error: str) -> None:
    row.status = "skipped"
    row.errors = [*row.errors, error]


def order_exists(db: Session, restaurant_id: int, uber_order_number: str) -> bool:
    return db.scalar(
        select(ClaimOrder.id).where(
            ClaimOrder.restaurant_id == restaurant_id,
            ClaimOrder.uber_order_number == uber_order_number,
        )
    ) is not None


def parse_amount(value: str) -> Decimal | None:
    cleaned = value.strip().replace("\u00a0", "").replace(" ", "")
    if not cleaned:
        return None

    if "," in cleaned and "." in cleaned:
        decimal_separator = "," if cleaned.rfind(",") > cleaned.rfind(".") else "."
        thousands_separator = "." if decimal_separator == "," else ","
        cleaned = cleaned.replace(thousands_separator, "")
        if decimal_separator == ",":
            cleaned = cleaned.replace(",", ".")
    elif "," in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")

    try:
        return Decimal(cleaned).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def normalize_optional_date(
    canonical_data: dict[str, str],
    normalized: dict[str, Any],
    errors: list[str],
    field: str,
) -> None:
    value = clean_optional(canonical_data.get(field))
    if not value:
        return
    parsed_date = parse_date_value(value)
    if parsed_date is None:
        errors.append(f"invalid_{field}")
    else:
        normalized[field] = parsed_date.isoformat()


def normalize_optional_time(
    canonical_data: dict[str, str],
    normalized: dict[str, Any],
    errors: list[str],
    field: str,
) -> None:
    value = clean_optional(canonical_data.get(field))
    if not value:
        return
    parsed_time = parse_time_value(value)
    if parsed_time is None:
        errors.append(f"invalid_{field}")
    else:
        normalized[field] = parsed_time.isoformat()


def normalize_optional_bool(
    canonical_data: dict[str, str],
    normalized: dict[str, Any],
    warnings: list[str],
    field: str,
    *,
    default: bool,
) -> None:
    value = clean_optional(canonical_data.get(field))
    if value is None:
        normalized[field] = default
        warnings.append(f"{field}_defaulted_to_{str(default).lower()}")
        return
    parsed_bool = parse_bool_value(value)
    if parsed_bool is None:
        warnings.append(f"{field}_ignored_invalid_boolean")
        return
    normalized[field] = parsed_bool


def parse_date_value(value: str) -> date | None:
    for format_value in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, format_value).date()
        except ValueError:
            continue
    return None


def parse_time_value(value: str) -> time | None:
    for format_value in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(value, format_value).time()
        except ValueError:
            continue
    return None


def parse_bool_value(value: str) -> bool | None:
    normalized = normalize_header(value)
    if normalized in {"1", "true", "vrai", "yes", "oui", "y"}:
        return True
    if normalized in {"0", "false", "faux", "no", "non", "n"}:
        return False
    return None


def parse_confirm_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value))


def parse_confirm_time(value: Any) -> time | None:
    if value is None:
        return None
    if isinstance(value, time):
        return value
    return time.fromisoformat(str(value))


def stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.time() == time(0, 0):
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return str(value).strip()


def normalize_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value.strip().lower())
    without_accents = "".join(character for character in normalized if not unicodedata.combining(character))
    return without_accents.replace(" ", "_").replace("-", "_")


def clean_optional(value: str | None) -> str | None:
    if value is None:
        return None
    trimmed = value.strip()
    return trimmed if trimmed else None


def is_empty_raw_row(raw_data: dict[str, str]) -> bool:
    return all(not clean_optional(value) for value in raw_data.values())


def safe_original_filename(filename: str | None) -> str:
    if not filename:
        raise OrderImportError("Import filename is required")
    original_filename = Path(filename).name.strip()
    if not original_filename or original_filename in {".", ".."}:
        raise OrderImportError("Import filename is invalid")
    return original_filename
