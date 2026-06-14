import csv
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal
from hashlib import sha256
from io import BytesIO, StringIO
from mimetypes import guess_type
from pathlib import Path
from typing import Any
from uuid import uuid4

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.models import SmartImportPreviewBatch, SmartImportPreviewFile, User
from app.models.domain import utc_now
from app.services.audit import add_audit_log

ALLOWED_EXTENSIONS = {"csv", "xlsx", "pdf", "jpg", "jpeg", "png", "webp", "heic", "heif", "zip"}
SPREADSHEET_EXTENSIONS = {"csv", "xlsx"}
IMAGE_EXTENSIONS = {"jpg", "jpeg", "png", "webp", "heic", "heif"}

REPORT_TYPE_LABELS = {
    "orders_report": "Rapport commandes Uber",
    "payments_report": "Rapport paiements Uber",
    "adjustments_report": "Rapport ajustements Uber",
    "combined_report": "Rapport Uber detecte",
}

COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "uber_store_id": (
        "uber_store_id",
        "store_id",
        "store_uuid",
        "restaurant_uuid",
        "merchant_store_id",
        "store id",
        "store uuid",
        "id. du restaurant",
        "id du restaurant",
        "id. externe du restaurant",
        "id externe du restaurant",
        "external restaurant id",
    ),
    "uber_store_name": (
        "store_name",
        "restaurant_name",
        "merchant_name",
        "restaurant",
        "store",
        "nom du restaurant",
    ),
    "uber_order_id": (
        "uber_order_id",
        "order_id",
        "order_uuid",
        "workflow_uuid",
        "commande_uber",
        "numero_commande",
        "numéro_commande",
        "uber_id",
        "id. de la commande",
        "id de la commande",
        "id. du flux",
        "id du flux",
        "uuid du processus",
        "process uuid",
    ),
    "display_id": ("display_id", "visible_id", "short_order_id", "order_display_id"),
    "order_status": ("order_status", "status", "current_state", "state", "statut", "statut de la commande"),
    "placed_at": (
        "placed_at",
        "order_date",
        "date",
        "date_commande",
        "created_at",
        "date de la commande",
        "heure d'acceptation de la commande",
        "heure de la commande",
        "heure d'acceptation par le marchand",
    ),
    "canceled_at": ("canceled_at", "cancellation_date", "cancellation_time", "heure_annulation", "annulation"),
    "order_total_amount": (
        "order_total_amount",
        "total",
        "amount",
        "order_amount",
        "montant",
        "montant_commande",
        "ventes (tva incluse)",
        "montant total",
    ),
    "transaction_type": (
        "transaction_type",
        "type",
        "adjustment_type",
        "payment_type",
        "line_item_type",
        "remboursements",
        "ajustements lies a des erreurs de commande (tva incluse)",
        "ajustements liés à des erreurs de commande (tva incluse)",
        "probleme avec la commande",
        "problème avec la commande",
        "informations concernant le probleme lie a l'article",
        "informations concernant le problème lié à l'article",
        "articles incorrects",
        "personnalisations incorrectes",
    ),
    "transaction_date": (
        "transaction_date",
        "payout_date",
        "payment_date",
        "date_transaction",
        "date_paiement",
        "date de la commande",
        "heure du remboursement",
        "heure de la commande",
    ),
    "payout_reference": (
        "payout_reference",
        "payout_id",
        "payment_reference",
        "versement",
        "reference_paiement",
        "id. de reference du versement",
        "id. de référence du versement",
    ),
    "amount": (
        "amount",
        "montant",
        "total",
        "value",
        "net_amount",
        "payout_amount",
        "montant total",
        "remboursements",
        "ajustements lies a des erreurs de commande (tva incluse)",
        "ajustements liés à des erreurs de commande (tva incluse)",
        "remboursement pris en charge par le commercant",
        "remboursement pris en charge par le commerçant",
        "client rembourse",
        "client remboursé",
    ),
    "currency": ("currency", "devise", "code de devise"),
}

COLUMN_ALIASES["uber_store_id"] += (
    "merchant_uuid",
    "merchant_id",
    "merchant store id",
    "merchant store uuid",
    "restaurant id",
    "restaurant uuid",
)
COLUMN_ALIASES["uber_order_id"] += (
    "workflow uuid",
    "workflow id",
    "order uuid",
    "order id",
    "order workflow id",
    "order workflow uuid",
    "uuid du workflow",
    "id du workflow",
)
COLUMN_ALIASES["placed_at"] += (
    "order date",
    "refund date",
    "refunded at",
    "date du remboursement",
    "date de remboursement",
    "date remboursement",
)
COLUMN_ALIASES["transaction_date"] += (
    "order date",
    "refund date",
    "refunded at",
    "date du remboursement",
    "date de remboursement",
    "date remboursement",
)
COLUMN_ALIASES["transaction_type"] += (
    "issue",
    "problem",
    "reason",
    "refund reason",
    "order issue",
    "accuracy issue",
    "defect category",
    "motif",
    "motif remboursement",
    "motif du remboursement",
    "categorie probleme",
    "categorie du probleme",
)
COLUMN_ALIASES["amount"] += (
    "refund amount",
    "refunded amount",
    "amount refunded",
    "customer refund",
    "customer refund amount",
    "merchant refund amount",
    "merchant charged amount",
    "amount charged to merchant",
    "deduction amount",
    "montant remboursement",
    "montant du remboursement",
    "montant rembourse",
    "montant facture au commercant",
    "montant deduit",
    "montant debit",
)

ORDER_FIELDS = {"uber_order_id", "order_status", "placed_at", "order_total_amount", "canceled_at"}
PAYMENT_FIELDS = {"transaction_date", "payout_reference", "amount", "currency"}
ADJUSTMENT_MARKERS = {"adjustment", "ajustement", "chargeback", "refund", "remboursement", "deduction"}
OFFICIAL_UBER_ACCURACY_MARKERS = {
    "inaccurate orders",
    "inaccurate orders v3",
    "top inaccurate items",
    "order accuracy",
    "order accuracy analytics",
    "commandes incorrectes",
    "commande incorrecte",
    "articles incorrects",
    "article incorrect",
}
ROW_LEVEL_ACCURACY_FIELDS = {"uber_store_id", "uber_order_id", "transaction_date", "amount"}
GENERIC_SINGLE_COLUMN_FIELDS = {"amount", "order_total_amount", "placed_at", "order_status", "transaction_type"}
EVIDENCE_HINTS = {
    "receipt": ("ticket", "receipt", "recu", "reçu", "facture"),
    "cancellation_proof": ("annulation", "cancel", "cancellation"),
    "preparation_proof": ("preparation", "préparation", "prep", "cuisine"),
    "waste_photo": ("gaspillage", "waste", "jete", "jeté"),
    "uber_screenshot": ("uber", "capture", "screenshot"),
    "delivery_proof": ("livraison", "delivery", "livre", "livré"),
    "missing_item": ("article manquant", "missing item", "produit manquant"),
    "packaging_photo": ("sac", "bag", "packaging", "emballage"),
}


@dataclass(frozen=True)
class HeaderDetection:
    header_index: int
    headers: list[str]
    score: int
    detected_fields: list[str]


@dataclass(frozen=True)
class FileClassification:
    original_filename: str
    file_type: str
    detected_category: str
    detected_report_type: str | None
    detected_evidence_type: str | None
    detected_restaurant_name: str | None
    detected_date_from: date | None
    detected_date_to: date | None
    header_row_number: int | None
    skipped_preamble_rows: int
    confidence: Decimal
    recommended_action: str
    warnings: list[str]
    detected_columns: list[str]
    metadata_json: dict[str, Any] | None


async def create_smart_import_preview(
    db: Session,
    current_user: User,
    files: list[UploadFile],
) -> SmartImportPreviewBatch:
    if not files:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="At least one file is required")

    expires_at = utc_now() + timedelta(hours=get_settings().smart_import_preview_expiry_hours)
    batch = SmartImportPreviewBatch(
        uploaded_by_user_id=current_user.id,
        status="previewed",
        files_count=len(files),
        total_files=len(files),
        expires_at=expires_at,
    )
    db.add(batch)
    db.flush()

    previews: list[tuple[FileClassification, StoredPreviewFile]] = []
    for file in files:
        stored = await store_preview_upload(batch.id, file)
        previews.append((classify_uploaded_file(stored.original_filename, stored.content), stored))

    preview_files: list[SmartImportPreviewFile] = []
    for preview, stored in previews:
        preview_file = SmartImportPreviewFile(
            batch_id=batch.id,
            original_filename=preview.original_filename,
            file_type=preview.file_type,
            temp_storage_path=stored.relative_path,
            mime_type=stored.mime_type,
            file_size=stored.file_size,
            checksum_sha256=stored.checksum_sha256,
            detected_category=preview.detected_category,
            detected_report_type=preview.detected_report_type,
            detected_evidence_type=preview.detected_evidence_type,
            detected_restaurant_name=preview.detected_restaurant_name,
            detected_date_from=preview.detected_date_from,
            detected_date_to=preview.detected_date_to,
            header_row_number=preview.header_row_number,
            skipped_preamble_rows=preview.skipped_preamble_rows,
            confidence=preview.confidence,
            recommended_action=preview.recommended_action,
            status="previewed",
            warnings=preview.warnings,
            detected_columns=preview.detected_columns,
            metadata_json=preview.metadata_json,
        )
        db.add(preview_file)
        preview_files.append(preview_file)
    db.flush()
    mark_exact_duplicate_preview_files(db, current_user, preview_files)
    mark_historical_exact_duplicate_preview_files(db, current_user, preview_files)

    add_audit_log(
        db,
        entity_type="smart_import_preview_batch",
        entity_id=batch.id,
        action="preview_smart_import",
        user_id=current_user.id,
        new_value={
            "files_count": len(files),
            "categories": [preview.detected_category for preview, _stored in previews],
            "recommended_actions": [preview.recommended_action for preview, _stored in previews],
        },
    )
    db.commit()
    db.refresh(batch)
    return batch


def mark_exact_duplicate_preview_files(
    db: Session,
    current_user: User,
    preview_files: list[SmartImportPreviewFile],
) -> None:
    checksum_groups: dict[str, list[SmartImportPreviewFile]] = {}
    for preview_file in preview_files:
        if preview_file.checksum_sha256:
            checksum_groups.setdefault(preview_file.checksum_sha256, []).append(preview_file)

    for duplicates in checksum_groups.values():
        if len(duplicates) <= 1:
            continue
        canonical = choose_duplicate_canonical_file(duplicates)
        for duplicate in duplicates:
            if duplicate.id == canonical.id or duplicate.status == "routed":
                continue
            mark_preview_file_as_exact_duplicate(db, current_user, duplicate, canonical)


def mark_historical_exact_duplicate_preview_files(
    db: Session,
    current_user: User,
    preview_files: list[SmartImportPreviewFile],
) -> None:
    for preview_file in preview_files:
        if preview_file.status != "previewed" or not preview_file.checksum_sha256:
            continue
        historical_files = db.scalars(
            select(SmartImportPreviewFile)
            .join(SmartImportPreviewBatch)
            .where(
                SmartImportPreviewFile.id != preview_file.id,
                SmartImportPreviewFile.batch_id != preview_file.batch_id,
                SmartImportPreviewFile.checksum_sha256 == preview_file.checksum_sha256,
                SmartImportPreviewFile.status.in_(("previewed", "confirmed", "routed", "manual_review")),
                SmartImportPreviewBatch.status.notin_(("cancelled", "expired")),
                or_(
                    SmartImportPreviewFile.status.in_(("confirmed", "routed", "manual_review")),
                    SmartImportPreviewBatch.expires_at.is_(None),
                    SmartImportPreviewBatch.expires_at > utc_now(),
                ),
            )
            .order_by(SmartImportPreviewFile.id)
        ).all()
        if not historical_files:
            continue
        canonical = choose_duplicate_canonical_file(historical_files)
        mark_preview_file_as_exact_duplicate(db, current_user, preview_file, canonical, scope="historical")


def choose_duplicate_canonical_file(preview_files: list[SmartImportPreviewFile]) -> SmartImportPreviewFile:
    action_rank = {"import_uber_reporting": 3, "import_evidence_bulk": 3, "manual_review": 1, "ignore": 0}

    def score(preview_file: SmartImportPreviewFile) -> tuple[int, Decimal, int, int]:
        name = preview_file.original_filename.lower()
        looks_like_copy = bool(re.search(r"\(\d+\)(?=\.[^.]+$|$)", name))
        return (
            action_rank.get(preview_file.recommended_action, 0),
            preview_file.confidence or Decimal("0"),
            0 if looks_like_copy else 1,
            -preview_file.id,
        )

    return max(preview_files, key=score)


def mark_preview_file_as_exact_duplicate(
    db: Session,
    current_user: User,
    duplicate: SmartImportPreviewFile,
    canonical: SmartImportPreviewFile,
    *,
    scope: str = "batch",
) -> None:
    if duplicate.error_message == f"exact_duplicate_of_file:{canonical.id}":
        return
    remove_preview_file_if_safe(duplicate)
    duplicate.status = "ignored"
    duplicate.recommended_action = "ignore"
    duplicate.destination_type = "duplicate_ignored"
    duplicate.destination_id = canonical.id
    duplicate.destination_url = f"/smart-import?preview={duplicate.batch_id}"
    duplicate.error_message = f"exact_duplicate_of_file:{canonical.id}"
    duplicate.warnings = sorted({*(duplicate.warnings or []), "exact_duplicate_ignored"})
    metadata = dict(duplicate.metadata_json or {})
    metadata.update(
        {
            "duplicate_match": "sha256",
            "duplicate_scope": scope,
            "duplicate_of_file_id": canonical.id,
            "duplicate_of_filename": canonical.original_filename,
        }
    )
    duplicate.metadata_json = metadata
    add_audit_log(
        db,
        entity_type="smart_import_preview_file",
        entity_id=duplicate.id,
        action="smart_import_file.duplicate_ignored",
        user_id=current_user.id,
        old_value={"filename": duplicate.original_filename, "checksum_sha256": duplicate.checksum_sha256},
        new_value={
            "canonical_file_id": canonical.id,
            "canonical_filename": canonical.original_filename,
            "match": "sha256",
            "scope": scope,
        },
    )


def remove_preview_file_if_safe(preview_file: SmartImportPreviewFile) -> None:
    if not preview_file.temp_storage_path:
        return
    root = ensure_smart_import_storage_root().resolve()
    target = (root / preview_file.temp_storage_path).resolve()
    if not target.is_relative_to(root):
        return
    if target.exists() and target.is_file():
        target.unlink()
    preview_file.temp_storage_path = None


@dataclass(frozen=True)
class StoredPreviewFile:
    original_filename: str
    content: bytes
    relative_path: str
    mime_type: str | None
    file_size: int
    checksum_sha256: str


async def store_preview_upload(batch_id: int, file: UploadFile) -> StoredPreviewFile:
    filename = safe_original_filename(file.filename or "fichier")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Smart import file cannot be empty")
    max_size = get_settings().import_max_file_size_mb * 1024 * 1024
    if len(content) > max_size:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="Smart import file is too large")
    suffix = file_extension(filename)
    if suffix not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported smart import file type")

    storage_root = ensure_smart_import_storage_root()
    relative_dir = Path("smart_import") / "previews" / f"batch_{batch_id}"
    target_dir = storage_root / relative_dir
    target_dir.mkdir(parents=True, exist_ok=True)
    extension = Path(filename).suffix.lower()
    internal_filename = f"{uuid4().hex}{extension}"
    relative_path = relative_dir / internal_filename
    target_path = storage_root / relative_path
    target_path.write_bytes(content)
    digest = sha256(content).hexdigest()
    return StoredPreviewFile(
        original_filename=filename,
        content=content,
        relative_path=relative_path.as_posix(),
        mime_type=file.content_type or guess_type(filename)[0],
        file_size=len(content),
        checksum_sha256=digest,
    )


def ensure_smart_import_storage_root() -> Path:
    root = get_settings().import_storage_dir
    root.mkdir(parents=True, exist_ok=True)
    return root


def resolve_preview_file_path(preview_file: SmartImportPreviewFile) -> Path:
    if not preview_file.temp_storage_path:
        raise HTTPException(status_code=status.HTTP_410_GONE, detail="Smart import preview file is not available")
    root = ensure_smart_import_storage_root().resolve()
    target = (root / preview_file.temp_storage_path).resolve()
    if not target.is_relative_to(root):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Smart import preview storage path is invalid")
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=status.HTTP_410_GONE, detail="Smart import preview file has expired")
    return target


def safe_original_filename(filename: str) -> str:
    original = Path(filename).name.strip()
    if not original or original in {".", ".."}:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Filename is invalid")
    return original


def confirm_smart_import_preview(
    db: Session,
    current_user: User,
    batch: SmartImportPreviewBatch,
) -> SmartImportPreviewBatch:
    if batch.status == "confirmed":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Smart import preview already confirmed")
    if batch.uploaded_by_user_id != current_user.id and current_user.role != "owner":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Smart import preview access denied")
    batch.status = "confirmed"
    batch.confirmed_at = utc_now()
    add_audit_log(
        db,
        entity_type="smart_import_preview_batch",
        entity_id=batch.id,
        action="confirm_smart_import_preview",
        user_id=current_user.id,
        new_value={"status": "confirmed"},
    )
    db.commit()
    db.refresh(batch)
    return batch


def classify_uploaded_file(filename: str, content: bytes) -> FileClassification:
    suffix = file_extension(filename)
    if suffix not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported smart import file type")
    if suffix in SPREADSHEET_EXTENSIONS:
        return classify_spreadsheet(filename, suffix, content)
    if suffix == "zip":
        return base_file_classification(
            filename=filename,
            suffix=suffix,
            category="zip",
            recommended_action="import_evidence_bulk",
            confidence=Decimal("0.80"),
            warnings=[],
            metadata={"message": "Archive ZIP detectee pour import massif de preuves."},
        )
    if suffix in IMAGE_EXTENSIONS or suffix == "pdf":
        evidence_type, confidence = detect_evidence_type_from_name(filename)
        return base_file_classification(
            filename=filename,
            suffix=suffix,
            category="evidence",
            evidence_type=evidence_type,
            recommended_action="import_evidence_bulk",
            confidence=confidence,
            warnings=["classification_limited_without_ocr"] if evidence_type == "other" else [],
        )
    return base_file_classification(filename, suffix, "unknown", "manual_review", Decimal("0.20"), ["unknown_file_category"])


def classify_spreadsheet(filename: str, suffix: str, content: bytes) -> FileClassification:
    rows = read_tabular_rows(content, suffix)
    warnings: list[str] = []
    if not rows:
        if looks_like_official_uber_accuracy_document(filename, []):
            return official_uber_accuracy_report_classification(filename, suffix, ["empty_official_uber_document"])
        return base_file_classification(filename, suffix, "unknown", "manual_review", Decimal("0.10"), ["empty_file"])

    header_detection = detect_best_header_row(rows)
    if not header_detection.detected_fields:
        if looks_like_official_uber_accuracy_document(filename, rows):
            return official_uber_accuracy_report_classification(
                filename,
                suffix,
                ["official_uber_analytics_document"],
                headers=header_detection.headers,
                header_index=header_detection.header_index,
                detected_fields=header_detection.detected_fields,
                rows_recognized=len([row for row in rows[header_detection.header_index + 1 :] if any(cell not in {None, ""} for cell in row)]),
            )
        return base_file_classification(
            filename,
            suffix,
            "unknown",
            "manual_review",
            Decimal("0.20"),
            ["no_known_columns_detected"],
        )

    data_rows = rows[header_detection.header_index + 1 :]
    if is_aggregate_uber_accuracy_document(filename, header_detection):
        return official_uber_accuracy_report_classification(
            filename,
            suffix,
            ["official_uber_aggregate_document"],
            headers=header_detection.headers,
            header_index=header_detection.header_index,
            detected_fields=header_detection.detected_fields,
            rows_recognized=len([row for row in data_rows if any(cell not in {None, ""} for cell in row)]),
        )
    report_type = detect_report_type(header_detection.detected_fields, data_rows, header_detection.headers)
    restaurant = detect_restaurant_name(header_detection.headers, data_rows)
    date_from, date_to = detect_period(header_detection.headers, data_rows)
    confidence = spreadsheet_confidence(header_detection, report_type, len(data_rows))
    if header_detection.header_index > 0:
        warnings.append("preamble_rows_ignored")
    if report_type == "combined_report":
        warnings.append("combined_report_detected")
    if looks_like_official_uber_accuracy_document(filename, [header_detection.headers, *data_rows[:3]]):
        warnings.append("official_uber_accuracy_report")

    return FileClassification(
        original_filename=filename,
        file_type=suffix,
        detected_category="uber_reporting",
        detected_report_type=report_type,
        detected_evidence_type=None,
        detected_restaurant_name=restaurant,
        detected_date_from=date_from,
        detected_date_to=date_to,
        header_row_number=header_detection.header_index + 1,
        skipped_preamble_rows=header_detection.header_index,
        confidence=confidence,
        recommended_action="import_uber_reporting" if confidence >= Decimal("0.45") else "manual_review",
        warnings=warnings,
        detected_columns=header_detection.headers,
        metadata_json={
            "detected_fields": header_detection.detected_fields,
            "rows_recognized": len([row for row in data_rows if any(cell not in {None, ""} for cell in row)]),
            "report_label": REPORT_TYPE_LABELS.get(report_type, "Rapport Uber detecte"),
        },
    )


def detect_best_header_row(rows: list[list[Any]]) -> HeaderDetection:
    candidates: list[HeaderDetection] = []
    for index, row in enumerate(rows[:5]):
        headers = [clean_header(value) for value in row]
        fields = detected_fields_for_headers(headers)
        unique_headers = len({header for header in headers if header})
        duplicate_penalty = max(0, len([header for header in headers if header]) - unique_headers)
        field_score = 0 if len(fields) == 1 and fields[0] in GENERIC_SINGLE_COLUMN_FIELDS else len(fields) * 3
        score = field_score + unique_headers - duplicate_penalty
        candidates.append(HeaderDetection(index, headers, score, fields))
    return max(candidates, key=lambda item: (item.score, -item.header_index))


def rows_to_dicts_with_detected_header(rows: list[list[Any]]) -> tuple[list[dict[str, Any]], HeaderDetection]:
    header_detection = detect_best_header_row(rows)
    data_rows = rows[header_detection.header_index + 1 :]
    headers = dedupe_headers(header_detection.headers)
    return [dict(zip(headers, row, strict=False)) for row in data_rows], header_detection


def read_tabular_rows(content: bytes, suffix: str) -> list[list[Any]]:
    if suffix == "csv":
        text = decode_csv(content)
        return list(csv.reader(StringIO(text)))
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    return [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]


def decode_csv(content: bytes) -> str:
    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError:
        return content.decode("latin-1")


def detected_fields_for_headers(headers: list[str]) -> list[str]:
    fields: list[str] = []
    normalized_headers = {normalize_for_match(header) for header in headers}
    for field, aliases in COLUMN_ALIASES.items():
        if any(normalize_for_match(alias) in normalized_headers for alias in aliases):
            fields.append(field)
    return sorted(set(fields))


def detect_report_type(fields: list[str], data_rows: list[list[Any]], headers: list[str]) -> str:
    field_set = set(fields)
    header_text = " ".join(headers)
    sample_text = " ".join(str(cell or "") for row in data_rows[:10] for cell in row)
    normalized_text = normalize_for_match(f"{header_text} {sample_text}")
    if looks_like_official_uber_accuracy_text(normalized_text) and ROW_LEVEL_ACCURACY_FIELDS.issubset(field_set):
        return "adjustments_report"
    has_order = bool(field_set & ORDER_FIELDS)
    has_payment = bool(field_set & PAYMENT_FIELDS)
    has_adjustment_marker = any(marker in normalized_text for marker in ADJUSTMENT_MARKERS)
    if has_order and (has_payment or has_adjustment_marker):
        return "combined_report"
    if has_adjustment_marker:
        return "adjustments_report"
    if has_payment and not has_order:
        return "payments_report"
    return "orders_report"


def detect_restaurant_name(headers: list[str], data_rows: list[list[Any]]) -> str | None:
    index = column_index(headers, "uber_store_name")
    if index is None:
        return None
    for row in data_rows[:25]:
        if index < len(row) and row[index] not in {None, ""}:
            return str(row[index]).strip()
    return None


def detect_period(headers: list[str], data_rows: list[list[Any]]) -> tuple[date | None, date | None]:
    indexes = [index for field in ("placed_at", "transaction_date", "canceled_at") if (index := column_index(headers, field)) is not None]
    dates: list[date] = []
    for row in data_rows[:1000]:
        for index in indexes:
            if index < len(row):
                parsed = parse_date_value(row[index])
                if parsed:
                    dates.append(parsed)
    if not dates:
        return None, None
    return min(dates), max(dates)


def column_index(headers: list[str], field: str) -> int | None:
    aliases = {normalize_for_match(alias) for alias in COLUMN_ALIASES[field]}
    for index, header in enumerate(headers):
        if normalize_for_match(header) in aliases:
            return index
    return None


def spreadsheet_confidence(header_detection: HeaderDetection, report_type: str, row_count: int) -> Decimal:
    score = min(0.95, 0.25 + (len(header_detection.detected_fields) * 0.08) + (0.15 if row_count > 0 else 0))
    if report_type == "combined_report":
        score += 0.04
    return Decimal(str(min(score, 0.98))).quantize(Decimal("0.01"))


def detect_evidence_type_from_name(filename: str) -> tuple[str, Decimal]:
    normalized = normalize_for_match(filename)
    for evidence_type, hints in EVIDENCE_HINTS.items():
        if any(normalize_for_match(hint) in normalized for hint in hints):
            return evidence_type, Decimal("0.72")
    return "other", Decimal("0.45")


def looks_like_official_uber_accuracy_document(filename: str, rows: list[list[Any]]) -> bool:
    sample_text = " ".join([filename, *(" ".join(str(cell or "") for cell in row) for row in rows[:8])])
    return looks_like_official_uber_accuracy_text(normalize_for_match(sample_text))


def looks_like_official_uber_accuracy_text(normalized_text: str) -> bool:
    return any(normalize_for_match(marker) in normalized_text for marker in OFFICIAL_UBER_ACCURACY_MARKERS)


def is_aggregate_uber_accuracy_document(filename: str, header_detection: HeaderDetection) -> bool:
    if not looks_like_official_uber_accuracy_document(filename, [header_detection.headers]):
        return False
    return not ROW_LEVEL_ACCURACY_FIELDS.issubset(set(header_detection.detected_fields))


def official_uber_accuracy_report_classification(
    filename: str,
    suffix: str,
    warnings: list[str],
    *,
    headers: list[str] | None = None,
    header_index: int | None = None,
    detected_fields: list[str] | None = None,
    rows_recognized: int = 0,
) -> FileClassification:
    fields = detected_fields or []
    has_row_level_transactions = ROW_LEVEL_ACCURACY_FIELDS.issubset(set(fields))
    effective_warnings = list(warnings)
    if not has_row_level_transactions:
        effective_warnings.append("missing_order_level_columns")
    return FileClassification(
        original_filename=filename,
        file_type=suffix,
        detected_category="uber_reporting",
        detected_report_type="adjustments_report",
        detected_evidence_type=None,
        detected_restaurant_name=None,
        detected_date_from=None,
        detected_date_to=None,
        header_row_number=(header_index + 1) if header_index is not None else None,
        skipped_preamble_rows=header_index or 0,
        confidence=Decimal("0.74") if has_row_level_transactions else Decimal("0.56"),
        recommended_action="import_uber_reporting",
        warnings=effective_warnings,
        detected_columns=headers or [],
        metadata_json={
            "official_uber_document": True,
            "row_level_transactions_available": has_row_level_transactions,
            "report_label": "Rapport officiel Uber a exploiter",
            "next_action": (
                "transactions_deductibles"
                if has_row_level_transactions
                else "completer_avec_export_detaille_commandes_et_montants"
            ),
            "rows_recognized": rows_recognized,
            "detected_columns": headers or [],
        },
    )


def parse_date_value(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def dedupe_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []
    for index, header in enumerate(headers, start=1):
        name = header or f"column_{index}"
        seen[name] = seen.get(name, 0) + 1
        result.append(name if seen[name] == 1 else f"{name}_{seen[name]}")
    return result


def clean_header(value: object) -> str:
    return str(value or "").strip()


def normalize_for_match(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").strip().lower())
    text = "".join(character for character in text if not unicodedata.combining(character))
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", text)).strip()


def file_extension(filename: str) -> str:
    return filename.lower().rsplit(".", 1)[-1] if "." in filename else ""


def base_file_classification(
    filename: str,
    suffix: str,
    category: str,
    recommended_action: str,
    confidence: Decimal,
    warnings: list[str],
    *,
    evidence_type: str | None = None,
    metadata: dict[str, Any] | None = None,
) -> FileClassification:
    return FileClassification(
        original_filename=filename,
        file_type=suffix,
        detected_category=category,
        detected_report_type=None,
        detected_evidence_type=evidence_type,
        detected_restaurant_name=None,
        detected_date_from=None,
        detected_date_to=None,
        header_row_number=None,
        skipped_preamble_rows=0,
        confidence=confidence,
        recommended_action=recommended_action,
        warnings=warnings,
        detected_columns=[],
        metadata_json=metadata,
    )
