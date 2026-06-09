import csv
from collections.abc import Iterable
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Any

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.auth import can_access_restaurant
from app.models import Restaurant, UberFinancialTransaction, UberOrderSnapshot, UberStoreMapping, User
from app.services.audit import add_audit_log


MAX_ROWS = 5000


async def import_uber_reporting_file(db: Session, current_user: User, file: UploadFile) -> dict[str, object]:
    filename = file.filename or "uber-report"
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix not in {"csv", "xlsx"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only CSV and XLSX reports are supported")

    content = await file.read()
    rows = parse_rows(content, suffix)
    snapshots_created = 0
    transactions_created = 0
    rows_skipped = 0
    errors: list[str] = []

    for row_number, raw_row in enumerate(rows, start=2):
        if row_number > MAX_ROWS + 1:
            errors.append("Maximum import row limit reached")
            break
        row = normalize_keys(raw_row)
        if not any(value not in {None, ""} for value in row.values()):
            rows_skipped += 1
            continue
        mapping = resolve_mapping(db, row)
        if mapping is None:
            rows_skipped += 1
            errors.append(f"Row {row_number}: unknown uber_store_id")
            continue
        if not can_access_restaurant(db, current_user, mapping.restaurant_id):
            rows_skipped += 1
            errors.append(f"Row {row_number}: restaurant access denied")
            continue

        if is_transaction_row(row):
            transaction = build_transaction(mapping, row)
            db.add(transaction)
            transactions_created += 1
        else:
            snapshot = upsert_snapshot(db, mapping, row)
            if snapshot:
                snapshots_created += 1

    add_audit_log(
        db,
        entity_type="uber_reporting_import",
        entity_id=current_user.id,
        action="import_uber_reporting",
        user_id=current_user.id,
        new_value={
            "snapshots_created": snapshots_created,
            "transactions_created": transactions_created,
            "rows_skipped": rows_skipped,
        },
    )
    db.commit()
    return {
        "snapshots_created": snapshots_created,
        "transactions_created": transactions_created,
        "rows_skipped": rows_skipped,
        "errors": errors,
    }


def parse_rows(content: bytes, suffix: str) -> list[dict[str, Any]]:
    if suffix == "csv":
        text = content.decode("utf-8-sig")
        return list(csv.DictReader(StringIO(text)))
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows_iter)]
    return [dict(zip(headers, values, strict=False)) for values in rows_iter]


def normalize_keys(row: dict[str, Any]) -> dict[str, Any]:
    return {str(key).strip().lower(): value for key, value in row.items() if key is not None}


def resolve_mapping(db: Session, row: dict[str, Any]) -> UberStoreMapping | None:
    store_id = text_value(row, "uber_store_id", "store_id", "restaurant_store_id")
    if store_id:
        return db.scalar(select(UberStoreMapping).where(UberStoreMapping.uber_store_id == store_id))
    restaurant_id = int_value(row, "restaurant_id")
    if restaurant_id:
        restaurant = db.get(Restaurant, restaurant_id)
        if restaurant is None:
            return None
        return db.scalar(select(UberStoreMapping).where(UberStoreMapping.restaurant_id == restaurant_id))
    return None


def is_transaction_row(row: dict[str, Any]) -> bool:
    return bool(text_value(row, "transaction_type", "type_transaction", "financial_type"))


def build_transaction(mapping: UberStoreMapping, row: dict[str, Any]) -> UberFinancialTransaction:
    amount = decimal_value(row, "amount", "transaction_amount", "montant") or Decimal("0")
    return UberFinancialTransaction(
        restaurant_id=mapping.restaurant_id,
        uber_store_id=mapping.uber_store_id,
        uber_order_id=text_value(row, "uber_order_id", "order_id", "uber_order_number"),
        transaction_type=text_value(row, "transaction_type", "type_transaction", "financial_type") or "unknown",
        amount=amount,
        currency=text_value(row, "currency", "devise") or "EUR",
        transaction_date=date_value(row, "transaction_date", "date") or date.today(),
        payout_reference=text_value(row, "payout_reference", "payout_id"),
        raw_payload_json=safe_json(row),
        imported_from="manager_export",
    )


def upsert_snapshot(db: Session, mapping: UberStoreMapping, row: dict[str, Any]) -> UberOrderSnapshot | None:
    uber_order_id = text_value(row, "uber_order_id", "order_id", "uber_order_number")
    if not uber_order_id:
        return None
    snapshot = db.scalar(
        select(UberOrderSnapshot).where(
            UberOrderSnapshot.uber_store_id == mapping.uber_store_id,
            UberOrderSnapshot.uber_order_id == uber_order_id,
        )
    )
    if snapshot is None:
        snapshot = UberOrderSnapshot(
            restaurant_id=mapping.restaurant_id,
            uber_store_id=mapping.uber_store_id,
            uber_order_id=uber_order_id,
            raw_payload_json=safe_json(row),
            imported_from="manager_export",
            current_state="unknown",
        )
        db.add(snapshot)
        db.flush()

    snapshot.display_id = text_value(row, "display_id", "display_order_id")
    snapshot.current_state = text_value(row, "current_state", "state", "status") or "unknown"
    snapshot.placed_at = datetime_value(row, "placed_at", "order_date")
    snapshot.canceled_at = datetime_value(row, "canceled_at", "cancelled_at", "cancellation_date")
    snapshot.order_total_amount = decimal_value(row, "order_total_amount", "order_amount", "total")
    snapshot.currency = text_value(row, "currency", "devise") or "EUR"
    snapshot.raw_payload_json = safe_json(row)
    return snapshot


def safe_json(row: dict[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in row.items():
        if isinstance(value, datetime):
            result[key] = value.isoformat()
        elif isinstance(value, date):
            result[key] = value.isoformat()
        elif isinstance(value, Decimal):
            result[key] = str(value)
        else:
            result[key] = value
    return result


def text_value(row: dict[str, Any], *keys: str) -> str | None:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return None


def int_value(row: dict[str, Any], *keys: str) -> int | None:
    value = text_value(row, *keys)
    if value is None:
        return None
    try:
        return int(value)
    except ValueError:
        return None


def decimal_value(row: dict[str, Any], *keys: str) -> Decimal | None:
    value = text_value(row, *keys)
    if value is None:
        return None
    normalized = value.replace(" ", "").replace(",", ".")
    try:
        return Decimal(normalized)
    except InvalidOperation:
        return None


def date_value(row: dict[str, Any], *keys: str) -> date | None:
    value = first_value(row, keys)
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip() if value is not None else ""
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def datetime_value(row: dict[str, Any], *keys: str) -> datetime | None:
    value = first_value(row, keys)
    if isinstance(value, datetime):
        return value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value
    parsed_date = date_value(row, *keys)
    if parsed_date:
        return datetime.combine(parsed_date, datetime.min.time(), tzinfo=timezone.utc)
    return None


def first_value(row: dict[str, Any], keys: Iterable[str]) -> Any:
    for key in keys:
        value = row.get(key)
        if value not in {None, ""}:
            return value
    return None
