from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO

from fastapi import HTTPException, status
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.models import ClaimOrder, EvidenceAnalysisResult, EvidenceImportBatch, EvidenceImportedFile, Restaurant, User
from app.models.domain import utc_now
from app.services.audit import add_audit_log
from app.services.bulk_evidence_import_service import BulkEvidenceImportError, resolve_imported_file_path

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".tif", ".tiff", ".bmp"}
PDF_EXTENSION = ".pdf"
EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}

ORDER_PATTERN = re.compile(r"\b(UBER(?:[-_\s]?[A-Z0-9]+){1,5})\b", re.IGNORECASE)
UUID_TEXT_PATTERN = re.compile(
    r"\b[a-f0-9]{8}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{12}\b",
    re.IGNORECASE,
)
INVALID_ORDER_TOKENS = {
    "A",
    "AU",
    "AUX",
    "CE",
    "CES",
    "DE",
    "DES",
    "DU",
    "EATS",
    "GUIDE",
    "POUR",
    "RAPPORT",
    "SUR",
    "VOTRE",
}
CONTEXTUAL_DISPLAY_PATTERN = re.compile(
    r"(?:n(?:umero)?\s*(?:de)?\s*commande|id\s*(?:de)?\s*(?:la)?\s*commande|"
    r"commande|cmd|order\s*(?:id)?|ticket|receipt|uber\s*(?:order)?\s*(?:id)?|id)[\s:_#-]{0,12}"
    r"([A-Z0-9][A-Z0-9-]{3,40})",
    re.IGNORECASE,
)
DISPLAY_PATTERN = re.compile(r"\b([A-Z0-9]{5,10})\b")
HASH_DISPLAY_PATTERN = re.compile(r"(?<![A-Z0-9])#\s*([A-Z0-9]{4,10})(?![A-Z0-9])", re.IGNORECASE)
AMOUNT_PATTERN = re.compile(r"(?<!\d)(\d{1,4}(?:[.,]\d{1,2}))(?!\d)")
LABELED_AMOUNT_PATTERN = re.compile(
    r"(?:montant\s*(?:total)?|total|amount|ventes|prix|remboursement|deduction|deduit)"
    r"[\s:=-]{0,12}(?:EUR|euros?|€)?\s*(\d{1,4}(?:[.,]\d{1,2})?)(?:\s*(?:EUR|euros?|€))?",
    re.IGNORECASE,
)
DATE_PATTERN = re.compile(r"\b(\d{4}[-/]\d{2}[-/]\d{2}|\d{2}[/-]\d{2}[/-]\d{4})\b")
CUSTOMER_LABEL_PATTERNS = (
    re.compile(
        r"(?:nom\s*(?:du)?\s*client|client|customer\s*(?:name)?|eater)\s*[:=-]\s*([^\n\r,;|]{2,80})",
        re.IGNORECASE,
    ),
    re.compile(
        r"(?:a|à)\s*pr[eé]par(?:er|e|é)?\s*pour\s+([^\n\r,;|]{2,80})",
        re.IGNORECASE,
    ),
    re.compile(r"prepared\s+for\s+([^\n\r,;|]{2,80})", re.IGNORECASE),
)
ADDITIONAL_CUSTOMER_LABEL_PATTERNS = (
    re.compile(
        r"(?:nom|name)\s*[:=-]\s*([^\n\r,;|]{2,120})",
        re.IGNORECASE,
    ),
    re.compile(
        r"(?:commande|order)\s+(?:pour|for)\s+([^\n\r,;|]{2,120})",
        re.IGNORECASE,
    ),
)
RESTAURANT_LABEL_PATTERN = re.compile(
    r"(?:nom\s*(?:du)?\s*restaurant|restaurant|store|merchant)\s*[:=-]\s*([^\n\r,;|]{2,100})",
    re.IGNORECASE,
)
MONTH_NAMES = {
    "janvier": 1,
    "jan": 1,
    "january": 1,
    "february": 2,
    "fevrier": 2,
    "février": 2,
    "feb": 2,
    "mars": 3,
    "march": 3,
    "mar": 3,
    "avril": 4,
    "april": 4,
    "apr": 4,
    "mai": 5,
    "may": 5,
    "juin": 6,
    "june": 6,
    "jun": 6,
    "juillet": 7,
    "july": 7,
    "jul": 7,
    "aout": 8,
    "août": 8,
    "august": 8,
    "aug": 8,
    "septembre": 9,
    "september": 9,
    "sep": 9,
    "octobre": 10,
    "october": 10,
    "oct": 10,
    "novembre": 11,
    "november": 11,
    "nov": 11,
    "decembre": 12,
    "décembre": 12,
    "december": 12,
    "dec": 12,
}
MONTH_NAME_PATTERN = re.compile(
    r"\b(?:(?:effectu[eé]e|pass[eé]e|created|ordered|date)\s*(?:le|on)?\s*)?"
    r"(?:(\d{1,2})\s+([A-Za-zÀ-ÿ]+)|([A-Za-zÀ-ÿ]+)\s+(\d{1,2}))"
    r"(?:[\s,]+(\d{4}))?\b",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class EvidenceAnalysisPayload:
    detected_evidence_type: str
    restaurant_name: str | None
    uber_order_number: str | None
    display_id: str | None
    order_date: date | None
    order_amount: Decimal | None
    currency: str | None
    customer_name: str | None
    keywords: list[str]
    classification_confidence: Decimal
    extraction_confidence: Decimal
    needs_manual_review: bool
    notes: str
    extracted_text: str


class EvidenceAIAnalysisService:
    def analyze_pending_batches(
        self,
        db: Session,
        current_user: User,
        *,
        restaurant_id: int | None,
        limit: int,
        batch_ids: list[int] | None = None,
        reanalyze_weak: bool = False,
    ) -> dict[str, object]:
        from app.services.evidence_matching_service import EvidenceMatchingService

        batches_statement = select(EvidenceImportBatch).order_by(EvidenceImportBatch.id.desc())
        if batch_ids is not None:
            if not batch_ids:
                return empty_analysis_result()
            batches_statement = batches_statement.where(EvidenceImportBatch.id.in_(batch_ids))
        else:
            batches_statement = batches_statement.limit(50)
        if restaurant_id is not None and batch_ids is None:
            batches_statement = batches_statement.where(EvidenceImportBatch.restaurant_id == restaurant_id)
        batches = db.scalars(batches_statement).all()
        analyzed_files_count = 0
        auto_matched_count = 0
        needs_review_count = 0
        failed_files_count = 0
        errors: list[str] = []
        remaining_limit = limit
        for batch in batches:
            if remaining_limit <= 0:
                break
            pending_files = [
                imported_file
                for imported_file in batch.files
                if imported_file.status in {"analysis_pending", "stored", "failed"}
            ]
            if pending_files:
                result = self.analyze_batch(db, current_user, batch, provider="fake", limit=remaining_limit)
                analyzed_files_count += int(result.get("analyzed_files_count", 0))
                auto_matched_count += int(result.get("auto_matched_count", 0))
                needs_review_count += int(result.get("needs_review_count", 0))
                failed_files_count += int(result.get("failed_files_count", 0))
                errors.extend(str(error) for error in result.get("errors", []))
                remaining_limit -= len(pending_files)

            matching_service = EvidenceMatchingService()
            if reanalyze_weak and remaining_limit > 0:
                weak_files = [
                    imported_file
                    for imported_file in batch.files
                    if imported_file.status == "analyzed"
                    and not has_attached_decision(imported_file)
                    and should_reanalyze_weak_file(imported_file)
                ]
                for imported_file in weak_files[:remaining_limit]:
                    try:
                        improved = self.reanalyze_file_if_improved(db, current_user, imported_file, matching_service)
                        if improved:
                            analyzed_files_count += 1
                            latest = latest_analysis_result(imported_file)
                            if latest is not None and latest.status == "manual_review":
                                needs_review_count += 1
                    except Exception as exc:  # noqa: BLE001 - one bad file must not stop the whole GO run.
                        failed_files_count += 1
                        errors.append(f"file {imported_file.id}: {exc}")
                    remaining_limit -= 1

            for imported_file in batch.files:
                if remaining_limit <= 0:
                    break
                if imported_file.status != "analyzed" or has_attached_decision(imported_file):
                    continue
                latest = latest_analysis_result(imported_file)
                if latest is None:
                    continue
                before_auto = len([candidate for candidate in imported_file.match_candidates if candidate.status == "auto_attached"])
                matching_service.create_candidates(db, current_user, imported_file, latest)
                after_auto = len([candidate for candidate in imported_file.match_candidates if candidate.status == "auto_attached"])
                auto_matched_count += max(after_auto - before_auto, 0)
                analyzed_files_count += 1
                remaining_limit -= 1

        for batch in batches:
            refresh_batch_analysis_counts(batch)
        return {
            "analyzed_files_count": analyzed_files_count,
            "auto_matched_count": auto_matched_count,
            "needs_review_count": needs_review_count,
            "failed_files_count": failed_files_count,
            "errors": errors,
        }

    def reanalyze_file_if_improved(
        self,
        db: Session,
        current_user: User,
        imported_file: EvidenceImportedFile,
        matching_service,
    ) -> bool:
        previous = latest_analysis_result(imported_file)
        if previous is None:
            return False
        payload = self.analyze_file(db, imported_file, provider="fake")
        if analysis_payload_score(payload) <= analysis_result_score(previous):
            return False
        result = build_analysis_result(imported_file, "fake", payload)
        db.add(result)
        db.flush()
        db.expire(imported_file, ["analysis_results"])
        candidates = matching_service.create_candidates(db, current_user, imported_file, result)
        result.matching_confidence = max((candidate.match_score for candidate in candidates), default=Decimal("0"))
        imported_file.status = "analyzed"
        imported_file.updated_at = utc_now()
        db.expire(imported_file, ["analysis_results", "match_candidates", "attachment_decisions"])
        add_audit_log(
            db,
            entity_type="evidence_imported_file",
            entity_id=imported_file.id,
            action="evidence_import_file.reanalyzed_with_stronger_identity",
            user_id=current_user.id,
            new_value={
                "previous_score": analysis_result_score(previous),
                "new_score": analysis_payload_score(payload),
                "detected_order": payload.uber_order_number or payload.display_id,
                "detected_restaurant": payload.restaurant_name,
                "detected_customer": payload.customer_name,
            },
        )
        return True

    def analyze_batch(
        self,
        db: Session,
        current_user: User,
        batch: EvidenceImportBatch,
        *,
        provider: str,
        limit: int,
    ) -> dict[str, object]:
        if provider == "openai_vision" and not get_settings().ai_evidence_analysis_enabled:
            raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="OpenAI evidence analysis is disabled")
        if provider == "local_ocr" and not get_settings().ocr_local_enabled:
            raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="Local OCR is disabled")

        batch.status = "analyzing"
        errors: list[str] = []
        files = db.scalars(
            select(EvidenceImportedFile)
            .where(
                EvidenceImportedFile.batch_id == batch.id,
                EvidenceImportedFile.status.in_(("analysis_pending", "stored", "failed")),
            )
            .order_by(EvidenceImportedFile.id)
            .limit(limit)
        ).all()

        from app.services.evidence_matching_service import EvidenceMatchingService

        matching_service = EvidenceMatchingService()
        for imported_file in files:
            try:
                payload = self.analyze_file(db, imported_file, provider)
                result = build_analysis_result(imported_file, provider, payload)
                db.add(result)
                db.flush()
                candidates = matching_service.create_candidates(db, current_user, imported_file, result)
                result.matching_confidence = max((candidate.match_score for candidate in candidates), default=Decimal("0"))
                imported_file.status = "analyzed"
                imported_file.updated_at = utc_now()
            except Exception as exc:
                imported_file.status = "failed"
                imported_file.updated_at = utc_now()
                errors.append(f"file {imported_file.id}: {exc}")

        refresh_batch_analysis_counts(batch)
        add_audit_log(
            db,
            entity_type="evidence_import_batch",
            entity_id=batch.id,
            action="evidence_import_batch.analyzed",
            user_id=current_user.id,
            new_value={
                "provider": provider,
                "analyzed_files_count": batch.analyzed_files_count,
                "auto_matched_count": batch.auto_matched_count,
                "needs_review_count": batch.needs_review_count,
                "failed_files_count": batch.failed_files_count,
                "errors": errors,
            },
        )
        return {
            "batch_id": batch.id,
            "status": batch.status,
            "analyzed_files_count": batch.analyzed_files_count,
            "auto_matched_count": batch.auto_matched_count,
            "needs_review_count": batch.needs_review_count,
            "failed_files_count": batch.failed_files_count,
            "errors": errors,
        }

    def analyze_file(self, db: Session, imported_file: EvidenceImportedFile, provider: str) -> EvidenceAnalysisPayload:
        if provider == "openai_vision":
            settings = get_settings()
            if not settings.ai_evidence_analysis_enabled or not settings.openai_api_key:
                raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="OpenAI evidence analysis is disabled")
        text = build_local_text(imported_file)
        evidence_type, keywords, classification_confidence = classify_evidence_type(text, imported_file.original_filename)
        order_number = extract_order_number(text)
        display_id = extract_display_id(text, order_number)
        amount = extract_amount(text)
        fallback_year = imported_file.created_at.year if imported_file.created_at else None
        detected_date = extract_date(text, fallback_year=fallback_year)
        restaurant_name = detect_restaurant_name(db, text, imported_file.batch.restaurant_id)
        customer_name = detect_customer_name(db, text, imported_file.batch.restaurant_id, restaurant_name)
        identity_score = sum(1 for value in (order_number, display_id, restaurant_name, customer_name, amount, detected_date) if value)
        extraction_confidence = Decimal("0.88") if order_number else Decimal("0.72") if display_id else Decimal("0.55")
        if restaurant_name and (order_number or display_id or customer_name):
            extraction_confidence = max(extraction_confidence, Decimal("0.82"))
        if restaurant_name and (order_number or display_id) and amount is not None:
            extraction_confidence = max(extraction_confidence, Decimal("0.90"))
        needs_manual_review = evidence_type == "unknown" or identity_score < 2
        return EvidenceAnalysisPayload(
            detected_evidence_type=evidence_type,
            restaurant_name=restaurant_name,
            uber_order_number=order_number,
            display_id=display_id,
            order_date=detected_date,
            order_amount=amount,
            currency="EUR" if amount is not None else None,
            customer_name=customer_name,
            keywords=keywords,
            classification_confidence=classification_confidence,
            extraction_confidence=extraction_confidence,
            needs_manual_review=needs_manual_review,
            notes=analysis_notes(evidence_type, order_number, display_id, restaurant_name, customer_name, amount),
            extracted_text=text,
        )


def build_local_text(imported_file: EvidenceImportedFile) -> str:
    filename_text = imported_file.original_filename.replace("_", " ").replace("-", " ")
    batch_context = ""
    if imported_file.batch.restaurant is not None:
        batch_context = imported_file.batch.restaurant.name
    try:
        path = resolve_imported_file_path(imported_file)
        content = path.read_bytes()
        decoded = extract_document_text(path.suffix.lower(), content)
    except (BulkEvidenceImportError, UnicodeDecodeError):
        decoded = ""
    return f"{filename_text}\n{batch_context}\n{decoded}".strip()


def build_analysis_result(
    imported_file: EvidenceImportedFile,
    provider: str,
    payload: EvidenceAnalysisPayload,
) -> EvidenceAnalysisResult:
    return EvidenceAnalysisResult(
        imported_file_id=imported_file.id,
        provider=provider,
        model_name=get_settings().openai_evidence_model if provider == "openai_vision" else None,
        status="manual_review" if payload.needs_manual_review else "success",
        extracted_text=payload.extracted_text,
        detected_evidence_type=payload.detected_evidence_type,
        detected_restaurant_name=payload.restaurant_name,
        detected_uber_order_number=payload.uber_order_number,
        detected_display_id=payload.display_id,
        detected_order_date=payload.order_date,
        detected_order_amount=payload.order_amount,
        detected_currency=payload.currency,
        detected_keywords_json=payload.keywords,
        classification_confidence=payload.classification_confidence,
        extraction_confidence=payload.extraction_confidence,
        matching_confidence=Decimal("0"),
        raw_result_json={
            "detected_evidence_type": payload.detected_evidence_type,
            "restaurant_name": payload.restaurant_name,
            "uber_order_number": payload.uber_order_number,
            "display_id": payload.display_id,
            "order_date": payload.order_date.isoformat() if payload.order_date else None,
            "order_amount": str(payload.order_amount) if payload.order_amount is not None else None,
            "currency": payload.currency,
            "customer_name": payload.customer_name,
            "keywords": payload.keywords,
            "classification_confidence": str(payload.classification_confidence),
            "extraction_confidence": str(payload.extraction_confidence),
            "needs_manual_review": payload.needs_manual_review,
            "notes": payload.notes,
            "unified_order_proof": looks_like_unified_order_proof(payload.extracted_text.lower()),
        },
    )


def extract_document_text(suffix: str, content: bytes) -> str:
    if suffix == PDF_EXTENSION:
        pieces = [decode_text_payload(content), extract_pdf_text(content)]
    elif suffix in IMAGE_EXTENSIONS:
        ocr_text = extract_image_ocr_text(content)
        decoded_text = decode_text_payload(content)
        pieces = [ocr_text or (decoded_text if looks_like_labeled_business_text(decoded_text) else "")]
    elif suffix in EXCEL_EXTENSIONS:
        pieces = [extract_excel_text(content)]
    else:
        pieces = [decode_text_payload(content)]
    return "\n".join(piece for piece in pieces if piece).strip()


def decode_text_payload(content: bytes) -> str:
    decoded = content[:32768].decode("utf-8", errors="ignore")
    readable = "".join(char if char.isprintable() or char in "\n\r\t" else " " for char in decoded)
    return readable.strip()


def looks_like_labeled_business_text(value: str) -> bool:
    normalized = normalize_for_match(value)
    label_hits = sum(
        1
        for token in (
            "restaurant",
            "client",
            "commande",
            "order",
            "montant",
            "remboursement",
            "annulation",
            "ticket",
        )
        if token in normalized
    )
    return label_hits >= 2


def extract_pdf_text(content: bytes) -> str:
    try:
        from pypdf import PdfReader

        reader = PdfReader(BytesIO(content))
        texts = [(page.extract_text() or "") for page in reader.pages[:5]]
        return "\n".join(texts).strip()
    except Exception:
        return ""


def extract_excel_text(content: bytes) -> str:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        pieces: list[str] = []
        for worksheet in workbook.worksheets[:8]:
            pieces.append(worksheet.title)
            header_values: list[str] | None = None
            for row in worksheet.iter_rows(min_row=1, max_row=250, max_col=60, values_only=True):
                values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                if values:
                    pieces.append(" | ".join(values))
                    if header_values is None and looks_like_header_row(values):
                        header_values = values
                    elif header_values is not None:
                        for header, value in zip(header_values, values, strict=False):
                            pieces.append(f"{header}: {value}")
        return "\n".join(pieces).strip()
    except Exception:
        return ""


def looks_like_header_row(values: list[str]) -> bool:
    normalized = [normalize_for_match(value) for value in values]
    header_tokens = {"restaurant", "client", "commande", "order", "montant", "total", "motif", "date"}
    return sum(1 for value in normalized if any(token in value for token in header_tokens)) >= 2


def extract_image_ocr_text(content: bytes) -> str:
    try:
        from PIL import Image, ImageOps
        import pytesseract

        with Image.open(BytesIO(content)) as image:
            normalized = ImageOps.exif_transpose(image)
            variants = build_ocr_image_variants(normalized)
            candidates: list[str] = []
            for variant in variants:
                for config in ("--psm 6", "--psm 4", "--psm 11", "--psm 12", ""):
                    try:
                        text = pytesseract.image_to_string(variant, lang="fra+eng", config=config).strip()
                    except pytesseract.TesseractError:
                        text = pytesseract.image_to_string(variant, config=config).strip()
                    if text:
                        candidates.append(text)
            return best_ocr_text(candidates)
    except Exception:
        return ""


def build_ocr_image_variants(image) -> list:
    from PIL import ImageFilter, ImageOps

    if image.mode not in {"L", "RGB"}:
        image = image.convert("RGB")
    base = image.convert("RGB")
    crops = [base]
    width, height = base.size
    if width >= 400 and height >= 400:
        crops.extend(
            [
                base.crop((0, 0, width, int(height * 0.42))),
                base.crop((0, int(height * 0.25), width, int(height * 0.75))),
                base.crop((0, int(height * 0.58), width, height)),
            ]
        )
        if width > height:
            crops.extend(
                [
                    base.crop((0, 0, int(width * 0.55), height)),
                    base.crop((int(width * 0.45), 0, width, height)),
                ]
            )
    variants = []
    for crop in crops:
        variants.extend(build_single_crop_ocr_variants(crop, ImageFilter, ImageOps))
    return variants


def build_single_crop_ocr_variants(image, image_filter, image_ops) -> list:
    grayscale = image_ops.grayscale(image)
    max_side = max(grayscale.size)
    scale = 2 if max_side < 2600 else 1
    if scale > 1:
        grayscale = grayscale.resize((grayscale.width * scale, grayscale.height * scale))
    contrasted = image_ops.autocontrast(grayscale)
    sharpened = contrasted.filter(image_filter.SHARPEN)
    sharpened_twice = sharpened.filter(image_filter.SHARPEN)
    threshold = contrasted.point(lambda pixel: 255 if pixel > 178 else 0)
    soft_threshold = contrasted.point(lambda pixel: 255 if pixel > 145 else 0)
    return [image, grayscale, contrasted, sharpened, sharpened_twice, threshold, soft_threshold]


def best_ocr_text(candidates: list[str]) -> str:
    if not candidates:
        return ""
    unique: dict[str, str] = {}
    for candidate in candidates:
        normalized = normalize_for_match(candidate)
        if normalized:
            unique.setdefault(normalized, candidate)
    if not unique:
        return max(candidates, key=len)

    def score(text: str) -> tuple[int, int]:
        normalized = normalize_for_match(text)
        business_hits = sum(
            1
            for token in (
                "commande",
                "order",
                "client",
                "restaurant",
                "total",
                "montant",
                "uber",
                "prepare",
                "remboursement",
                "annulation",
            )
            if token in normalized
        )
        return (business_hits, len(text))

    ranked = sorted(unique.values(), key=score, reverse=True)
    return merge_ocr_candidates(ranked[:4])


def merge_ocr_candidates(candidates: list[str]) -> str:
    lines: list[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        for line in candidate.splitlines():
            cleaned = re.sub(r"\s+", " ", line).strip()
            if not cleaned:
                continue
            normalized = normalize_for_match(cleaned)
            if normalized in seen:
                continue
            seen.add(normalized)
            lines.append(cleaned)
    return "\n".join(lines)


def classify_evidence_type(text: str, filename: str) -> tuple[str, list[str], Decimal]:
    lower = f"{filename} {text}".lower()
    rules = [
        (
            "uber_screenshot",
            (
                "order accuracy",
                "inaccurate orders",
                "top inaccurate items",
                "articles incorrects",
                "commandes incorrectes",
                "client rembourse",
                "remboursement pris en charge",
            ),
        ),
        ("cancellation_proof", ("annulation", "cancel", "canceled", "cancelled")),
        ("preparation_proof", ("preparation", "preparee", "préparée", "prepared", "ready")),
        ("waste_photo", ("waste", "gaspillage", "jetee", "jetée")),
        ("receipt", ("receipt", "ticket", "caisse", "recu", "reçu")),
        ("uber_screenshot", ("uber", "manager", "capture", "screenshot")),
        ("delivery_proof", ("delivery", "livraison", "delivered")),
        ("packaging_photo", ("packaging", "emballage", "pack")),
        ("sealed_bag_photo", ("sealed", "scelle", "scellé", "bag", "sac")),
        ("order_details_screenshot", ("order details", "details commande", "détails commande")),
        ("courier_statement", ("courier", "livreur", "coursier")),
        ("gps_or_route_proof", ("gps", "route", "trajet")),
        ("customer_contact_proof", ("customer contact", "client message", "whatsapp")),
    ]
    for evidence_type, keywords in rules:
        matched = [keyword for keyword in keywords if keyword in lower]
        if matched:
            return evidence_type, matched, Decimal("0.85")
    if looks_like_unified_order_proof(lower):
        return "receipt", ["ticket_commande_unique"], Decimal("0.80")
    return "unknown", [], Decimal("0.25")


def looks_like_unified_order_proof(normalized: str) -> bool:
    has_order_context = any(token in normalized for token in ("commande", "cmd", "order", "uber", "client"))
    has_preparation_context = any(token in normalized for token in ("prepare", "prepar", "emballe", "bag", "sac", "menu"))
    has_amount = AMOUNT_PATTERN.search(normalized) is not None
    return has_order_context and (has_preparation_context or has_amount)


def extract_order_number(text: str) -> str | None:
    match = ORDER_PATTERN.search(text)
    if match:
        candidate = re.sub(r"[-_\s]+", "-", match.group(1)).strip("-").upper()
        if valid_order_number(candidate):
            return candidate
    normalized_text = normalize_identifier_text(text)
    contextual_uuid = re.search(
        r"(?:commande|cmd|order|ticket|receipt|id)\s*[:#=-]{0,8}\s*"
        r"([a-f0-9]{8}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{12})",
        normalized_text,
        flags=re.IGNORECASE,
    )
    if contextual_uuid:
        return contextual_uuid.group(1).lower()
    if any(token in normalize_for_match(text) for token in ("commande", "order", "ticket uber", "receipt")):
        uuid_match = UUID_TEXT_PATTERN.search(text)
        if uuid_match:
            return uuid_match.group(0).lower()
    return None


def valid_order_number(value: str) -> bool:
    tokens = [token for token in re.split(r"[-_\s]+", value.upper()) if token]
    if len(tokens) < 2:
        return False
    if any(token in INVALID_ORDER_TOKENS for token in tokens[1:]):
        return False
    return True


def extract_display_id(text: str, order_number: str | None) -> str | None:
    if order_number:
        return None
    normalized_text = normalize_identifier_text(text)
    hash_match = HASH_DISPLAY_PATTERN.search(normalized_text)
    if hash_match:
        cleaned = clean_display_id(hash_match.group(1))
        if valid_display_id(cleaned):
            return f"#{cleaned}"
    match = CONTEXTUAL_DISPLAY_PATTERN.search(normalized_text)
    if match:
        cleaned = clean_display_id(match.group(1))
        if valid_display_id(cleaned):
            return cleaned
    if any(token in normalized_text.lower() for token in ("commande", "cmd", "order", "ticket", "receipt", "uber")):
        fallback = DISPLAY_PATTERN.search(normalized_text)
        if fallback:
            cleaned = clean_display_id(fallback.group(1))
            if valid_display_id(cleaned):
                return cleaned
    return None


def extract_amount(text: str) -> Decimal | None:
    cleaned_text = text.replace("\xa0", " ")
    labeled = LABELED_AMOUNT_PATTERN.search(cleaned_text)
    if labeled:
        if is_percent_match(cleaned_text, labeled):
            return None
        try:
            return Decimal(labeled.group(1).replace(",", ".")).quantize(Decimal("0.01"))
        except InvalidOperation:
            return None
    for match in AMOUNT_PATTERN.finditer(cleaned_text):
        if is_percent_match(cleaned_text, match):
            continue
        try:
            return Decimal(match.group(1).replace(",", ".")).quantize(Decimal("0.01"))
        except InvalidOperation:
            continue
    return None


def is_percent_match(text: str, match: re.Match[str]) -> bool:
    before = text[max(0, match.start() - 4) : match.start()]
    after = text[match.end() : match.end() + 4]
    return "%" in before or "%" in after


def extract_date(text: str, *, fallback_year: int | None = None) -> date | None:
    match = DATE_PATTERN.search(text)
    if match:
        raw = match.group(1)
        try:
            if raw[4:5] in {"-", "/"}:
                return date.fromisoformat(raw.replace("/", "-"))
            day, month, year = re.split(r"[/-]", raw)
            return date(int(year), int(month), int(day))
        except (ValueError, IndexError):
            pass
    for month_match in MONTH_NAME_PATTERN.finditer(text):
        if month_match.group(2):
            day_raw = month_match.group(1)
            month_raw = month_match.group(2)
        else:
            month_raw = month_match.group(3)
            day_raw = month_match.group(4)
        year_raw = month_match.group(5)
        month_number = MONTH_NAMES.get(month_raw.lower())
        year = int(year_raw) if year_raw else fallback_year
        if month_number is None or year is None:
            continue
        try:
            return date(year, month_number, int(day_raw))
        except ValueError:
            continue
    return None


def detect_restaurant_name(db: Session, text: str, batch_restaurant_id: int | None) -> str | None:
    if batch_restaurant_id is not None:
        restaurant = db.get(Restaurant, batch_restaurant_id)
        if restaurant is not None:
            return restaurant.name
    labeled = extract_labeled_restaurant_name(text)
    if labeled:
        resolved = resolve_restaurant_display_name(db, labeled)
        if resolved:
            return resolved
    normalized_text = normalize_for_match(text)
    restaurants = db.scalars(select(Restaurant).where(Restaurant.active.is_(True)).order_by(Restaurant.id)).all()
    matches: list[tuple[int, str]] = []
    for restaurant in restaurants:
        name = normalize_for_match(restaurant.name)
        if len(name) >= 4 and name in normalized_text:
            matches.append((len(name), restaurant.name))
    if not matches:
        return None
    matches.sort(reverse=True)
    return matches[0][1]


def detect_customer_name(db: Session, text: str, batch_restaurant_id: int | None, restaurant_name: str | None) -> str | None:
    labeled = extract_labeled_customer_name(text)
    if labeled:
        return labeled
    normalized_text = normalize_for_match(text)
    statement = select(ClaimOrder).where(ClaimOrder.customer_name.is_not(None)).order_by(ClaimOrder.id.desc()).limit(500)
    if batch_restaurant_id is not None:
        statement = statement.where(ClaimOrder.restaurant_id == batch_restaurant_id)
    elif restaurant_name:
        restaurant = db.scalar(select(Restaurant).where(Restaurant.name == restaurant_name))
        if restaurant is not None:
            statement = statement.where(ClaimOrder.restaurant_id == restaurant.id)
    candidates = []
    for order in db.scalars(statement).all():
        customer_name = clean_customer_name(order.customer_name or "")
        if not customer_name:
            continue
        normalized_name = normalize_for_match(customer_name)
        if len(normalized_name) >= 4 and normalized_name in normalized_text:
            candidates.append((len(normalized_name), customer_name))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


def normalize_for_match(value: str) -> str:
    without_accents = "".join(
        char for char in unicodedata.normalize("NFKD", value.lower()) if not unicodedata.combining(char)
    )
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", without_accents)).strip()


def normalize_identifier_text(value: str) -> str:
    without_accents = "".join(char for char in unicodedata.normalize("NFKD", value) if not unicodedata.combining(char))
    return without_accents.replace("_", " ")


def clean_display_id(value: str) -> str:
    cleaned = re.sub(r"[^A-Z0-9-]", "", value.upper()).strip("-")
    return cleaned[:20] if cleaned else ""


def valid_display_id(value: str | None) -> bool:
    if not value or len(value) < 4:
        return False
    if value in {"TICKET", "CLIENT", "ORDER", "COMMANDE", "RECEIPT"}:
        return False
    if value.isdigit():
        return False
    if re.fullmatch(r"20\d{6}", value):
        return False
    return any(char.isdigit() for char in value)


def extract_labeled_customer_name(text: str) -> str | None:
    for pattern in (*CUSTOMER_LABEL_PATTERNS, *ADDITIONAL_CUSTOMER_LABEL_PATTERNS):
        match = pattern.search(text)
        if not match:
            continue
        for value in match.groups():
            cleaned = clean_customer_name(value)
            if cleaned:
                return cleaned
    return None


def extract_labeled_restaurant_name(text: str) -> str | None:
    match = RESTAURANT_LABEL_PATTERN.search(text)
    if not match:
        return None
    return clean_human_label(match.group(1), max_length=100)


def clean_human_label(value: str, *, max_length: int) -> str | None:
    cleaned = trim_following_ocr_labels(re.sub(r"\s+", " ", value)).strip(" -_:;,.")
    if not cleaned or len(cleaned) < 2:
        return None
    blocked = {"commande", "order", "ticket", "receipt", "montant", "total"}
    if normalize_for_match(cleaned) in blocked:
        return None
    return cleaned[:max_length]


def trim_following_ocr_labels(value: str) -> str:
    parts = re.split(
        r"\s+(?=(?:restaurant|client|customer|eater|commande|cmd|order|date|montant|total|ticket|"
        r"receipt|remboursement|annulation|refund|cancel)\s*[:#=-])",
        value,
        maxsplit=1,
        flags=re.IGNORECASE,
    )
    return parts[0] if parts else value


def clean_customer_name(value: str) -> str | None:
    cleaned = clean_human_label(value, max_length=80)
    if cleaned is None:
        return None
    normalized = normalize_for_match(cleaned)
    if re.fullmatch(r"\d+(?:[.,]\d+)?\s*(?:EUR|euro|euros|%)?", cleaned, flags=re.IGNORECASE):
        return None
    if re.fullmatch(r"\d+(?:[.,]\d+)?", normalized):
        return None
    if re.fullmatch(r"\d+(?:[.,]\d+)?\s*(eur|euro|euros|percent|pourcent|%)?", normalized):
        return None
    blocked_labels = {
        "tva",
        "ht",
        "ttc",
        "eur",
        "euro",
        "euros",
        "nouveau client",
        "client rembourse",
        "client rembourse",
        "couverts jetables",
        "couverts jet",
        "montant total",
        "sous total",
        "frais de livraison",
    }
    if normalized in blocked_labels:
        return None
    letters = [char for char in cleaned if char.isalpha()]
    if len(letters) < 2:
        return None
    return cleaned


def resolve_restaurant_display_name(db: Session, value: str) -> str | None:
    from app.services.uber_reporting_import_service import resolve_mapping_by_store_name, resolve_restaurant_by_store_name

    mapping = resolve_mapping_by_store_name(db, value)
    if mapping is not None and mapping.restaurant is not None:
        return mapping.restaurant.name
    restaurant = resolve_restaurant_by_store_name(db, value)
    if restaurant is not None:
        return restaurant.name
    return None


def analysis_notes(
    evidence_type: str,
    order_number: str | None,
    display_id: str | None,
    restaurant_name: str | None,
    customer_name: str | None,
    amount: Decimal | None,
) -> str:
    missing = []
    if evidence_type == "unknown":
        missing.append("type de preuve")
    if not (order_number or display_id):
        missing.append("numero de commande")
    if not restaurant_name:
        missing.append("restaurant")
    if not customer_name:
        missing.append("client")
    if amount is None:
        missing.append("montant")
    if missing:
        return "Analyse deterministe sans OpenAI. Donnees encore manquantes: " + ", ".join(missing) + "."
    return "Analyse deterministe complete sans appel OpenAI reel."


def refresh_batch_analysis_counts(batch: EvidenceImportBatch) -> None:
    files = list(batch.files)
    analyzed = [item for item in files if item.status == "analyzed"]
    failed = [item for item in files if item.status == "failed"]
    batch.analyzed_files_count = len(analyzed)
    batch.failed_files_count = len(failed)
    batch.auto_matched_count = len(
        [
            candidate
            for imported_file in files
            for candidate in imported_file.match_candidates
            if candidate.status == "auto_attached"
        ]
    )
    batch.needs_review_count = len(
        [
            imported_file
            for imported_file in files
            if imported_file.status == "analyzed"
            and not any(candidate.status in {"accepted", "auto_attached"} for candidate in imported_file.match_candidates)
        ]
    )
    if failed and analyzed:
        batch.status = "partially_analyzed"
    elif failed and not analyzed:
        batch.status = "failed"
    elif analyzed:
        batch.status = "analyzed"
        batch.completed_at = utc_now()
    batch.updated_at = utc_now()


def latest_analysis_result(imported_file: EvidenceImportedFile) -> EvidenceAnalysisResult | None:
    if not imported_file.analysis_results:
        return None
    return sorted(imported_file.analysis_results, key=lambda item: item.id)[-1]


def should_reanalyze_weak_file(imported_file: EvidenceImportedFile) -> bool:
    latest = latest_analysis_result(imported_file)
    if latest is None:
        return False
    if latest.status not in {"manual_review", "success"}:
        return False
    if latest.raw_result_json and latest.raw_result_json.get("reanalysis_locked"):
        return False
    return analysis_result_score(latest) < 5


def analysis_payload_score(payload: EvidenceAnalysisPayload) -> int:
    return sum(
        1
        for value in (
            payload.uber_order_number or payload.display_id,
            payload.restaurant_name,
            payload.customer_name,
            payload.order_amount,
            payload.order_date,
            payload.detected_evidence_type != "unknown",
        )
        if value
    )


def analysis_result_score(result: EvidenceAnalysisResult) -> int:
    raw = result.raw_result_json or {}
    return sum(
        1
        for value in (
            result.detected_uber_order_number or result.detected_display_id,
            result.detected_restaurant_name,
            raw.get("customer_name"),
            result.detected_order_amount,
            result.detected_order_date,
            result.detected_evidence_type != "unknown",
        )
        if value
    )


def has_attached_decision(imported_file: EvidenceImportedFile) -> bool:
    return any(decision.decision == "attached" for decision in imported_file.attachment_decisions)


def empty_analysis_result() -> dict[str, object]:
    return {
        "analyzed_files_count": 0,
        "auto_matched_count": 0,
        "needs_review_count": 0,
        "failed_files_count": 0,
        "errors": [],
    }
