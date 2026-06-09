from datetime import timedelta
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import (
    EvidenceRequestTask,
    FollowUpTask,
    UberCustomerRefundDispute,
    UberReconciliationResult,
)
from app.models.domain import utc_now


def auth_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def create_restaurant(client: TestClient, name: str = "Recovery Restaurant") -> dict:
    response = client.post("/v1/restaurants", json={"name": name, "sender_email": "claims@example.com"})
    assert response.status_code == 201
    return response.json()


def create_user(client: TestClient, email: str, role: str) -> dict:
    response = client.post(
        "/v1/users",
        json={
            "email": email,
            "password": "user-password",
            "full_name": f"{role.title()} Recovery",
            "role": role,
            "active": True,
        },
    )
    assert response.status_code == 201
    return response.json()


def assign_restaurant(client: TestClient, user_id: int, restaurant_id: int) -> None:
    response = client.post(f"/v1/users/{user_id}/restaurants", json={"restaurant_id": restaurant_id})
    assert response.status_code == 201


def login(client: TestClient, email: str, password: str = "user-password") -> str:
    response = client.post("/v1/auth/login", json={"email": email, "password": password})
    assert response.status_code == 200
    return response.json()["access_token"]


def create_order(
    client: TestClient,
    restaurant_id: int,
    order_number: str,
    amount: str,
    *,
    status: str = "missing_evidence",
    recovered_amount: str | None = None,
) -> dict:
    response = client.post(
        "/v1/orders",
        json={
            "restaurant_id": restaurant_id,
            "uber_order_number": order_number,
            "order_amount": amount,
            "currency": "EUR",
            "accepted_by_restaurant": True,
            "prepared_before_cancellation": True,
            "status": status,
            "recovered_amount": recovered_amount,
        },
    )
    assert response.status_code == 201
    return response.json()


def add_reconciliation_result(
    db_session: Session,
    restaurant_id: int,
    order_id: str,
    *,
    status: str = "not_compensated",
    amount: str = "18.00",
) -> UberReconciliationResult:
    result = UberReconciliationResult(
        restaurant_id=restaurant_id,
        uber_order_id=order_id,
        display_id=order_id,
        status=status,
        reason="canceled_no_payment_found",
        order_amount=amount,
        paid_amount="0.00",
        refunded_amount="0.00",
        missing_amount=amount,
        currency="EUR",
        evidence_required=True,
    )
    db_session.add(result)
    db_session.commit()
    db_session.refresh(result)
    return result


def add_customer_refund(
    db_session: Session,
    restaurant_id: int,
    *,
    status: str = "needs_evidence",
    amount: str = "12.00",
    recovered_amount: str | None = None,
) -> UberCustomerRefundDispute:
    dispute = UberCustomerRefundDispute(
        restaurant_id=restaurant_id,
        uber_store_id="store-recovery",
        uber_order_id=f"UBER-RECOVERY-{status}",
        display_id=f"UBER-RECOVERY-{status}",
        dispute_type="customer_refund",
        reason="refund_without_sufficient_proof",
        status=status,
        customer_refund_amount=amount,
        recovered_amount=recovered_amount,
        order_amount=amount,
        currency="EUR",
        deducted_at=utc_now().date(),
        order_date=utc_now().date(),
        evidence_required=True,
        evidence_status="missing" if status == "needs_evidence" else "complete",
        raw_payload_json={"source": "test"},
    )
    db_session.add(dispute)
    db_session.commit()
    db_session.refresh(dispute)
    return dispute


def add_evidence_task(db_session: Session, order_id: int, restaurant_id: int) -> EvidenceRequestTask:
    task = EvidenceRequestTask(
        order_id=order_id,
        restaurant_id=restaurant_id,
        task_type="missing_receipt",
        required_evidence_type="receipt",
        status="pending",
        priority="high",
        title="Ticket requis",
        description="Preuve manquante",
        reason="test_recovery",
        due_at=utc_now() + timedelta(days=1),
    )
    db_session.add(task)
    db_session.commit()
    db_session.refresh(task)
    return task


def add_followup(db_session: Session, order_id: int) -> FollowUpTask:
    task = FollowUpTask(
        order_id=order_id,
        task_type="followup_1",
        status="pending",
        due_at=utc_now() - timedelta(hours=1),
    )
    db_session.add(task)
    db_session.commit()
    db_session.refresh(task)
    return task


def seed_recovery_data(client: TestClient, db_session: Session) -> dict:
    restaurant = create_restaurant(client, "Recovery A")
    other = create_restaurant(client, "Recovery B")
    missing_order = create_order(client, restaurant["id"], "UBER-RECOVERY-MISSING", "30.00", status="missing_evidence")
    sent_order = create_order(client, restaurant["id"], "UBER-RECOVERY-SENT", "20.00", status="sent")
    recovered_order = create_order(
        client,
        other["id"],
        "UBER-RECOVERY-PAID",
        "10.00",
        status="payment_confirmed",
        recovered_amount="10.00",
    )
    add_reconciliation_result(db_session, restaurant["id"], "UBER-RECOVERY-RECON", amount="18.00")
    add_customer_refund(db_session, restaurant["id"], status="needs_evidence", amount="12.00")
    add_customer_refund(db_session, restaurant["id"], status="payment_confirmed", amount="8.00", recovered_amount="8.00")
    add_customer_refund(db_session, restaurant["id"], status="refused", amount="5.00")
    add_evidence_task(db_session, missing_order["id"], restaurant["id"])
    add_followup(db_session, sent_order["id"])
    return {
        "restaurant": restaurant,
        "other": other,
        "missing_order": missing_order,
        "sent_order": sent_order,
        "recovered_order": recovered_order,
    }


def test_health_works(unauthenticated_client: TestClient) -> None:
    response = unauthenticated_client.get("/health")

    assert response.status_code == 200


def test_recovery_summary_calculates_core_totals(client: TestClient, db_session: Session) -> None:
    seed_recovery_data(client, db_session)

    response = client.get("/v1/recovery/summary")

    assert response.status_code == 200
    totals = response.json()["totals"]
    assert totals["detected_amount"] == "103.00"
    assert totals["claimable_amount"] == "80.00"
    assert totals["missing_evidence_amount"] == "60.00"
    assert totals["recovered_amount"] == "18.00"
    assert totals["refused_amount"] == "5.00"
    assert "recovery_rate" in totals


def test_recovery_cases_include_claims_reconciliation_and_customer_refunds(client: TestClient, db_session: Session) -> None:
    seed_recovery_data(client, db_session)

    response = client.get("/v1/recovery/cases")

    assert response.status_code == 200
    case_types = {case["case_type"] for case in response.json()["cases"]}
    assert {"claim_order", "reconciliation_result", "customer_refund_dispute"} <= case_types


def test_recovery_actions_include_evidence_and_followups(client: TestClient, db_session: Session) -> None:
    seed_recovery_data(client, db_session)

    response = client.get("/v1/recovery/actions")

    assert response.status_code == 200
    action_types = {action["action_type"] for action in response.json()["actions"]}
    assert "upload_evidence" in action_types
    assert "followup" in action_types


def test_manager_only_sees_assigned_restaurant(client: TestClient, db_session: Session) -> None:
    data = seed_recovery_data(client, db_session)
    manager = create_user(client, "manager-recovery@example.com", "manager")
    assign_restaurant(client, manager["id"], data["restaurant"]["id"])
    token = login(client, manager["email"])

    response = client.get("/v1/recovery/summary", headers=auth_headers(token))

    assert response.status_code == 200
    restaurants = {row["restaurant_name"] for row in response.json()["by_restaurant"]}
    assert restaurants == {"Recovery A"}


def test_staff_cannot_export_recovery(client: TestClient, db_session: Session) -> None:
    data = seed_recovery_data(client, db_session)
    staff = create_user(client, "staff-recovery@example.com", "staff")
    assign_restaurant(client, staff["id"], data["restaurant"]["id"])
    token = login(client, staff["email"])

    response = client.get("/v1/recovery/export/cases.csv", headers=auth_headers(token))

    assert response.status_code == 403


def test_recovery_summary_xlsx_export_works(client: TestClient, db_session: Session) -> None:
    seed_recovery_data(client, db_session)

    response = client.get("/v1/recovery/export/summary.xlsx")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    assert {"Summary", "By Restaurant", "By Category", "By Stage", "Top Recoverable", "Actions"} <= set(workbook.sheetnames)


def test_recovery_cases_csv_export_has_no_sensitive_paths(client: TestClient, db_session: Session) -> None:
    seed_recovery_data(client, db_session)

    response = client.get("/v1/recovery/export/cases.csv")

    assert response.status_code == 200
    content = response.text
    assert "storage_path" not in content
    assert "access_token" not in content
    assert "refresh_token" not in content


def test_recovery_case_filters(client: TestClient, db_session: Session) -> None:
    seed_recovery_data(client, db_session)

    response = client.get("/v1/recovery/cases?case_type=customer_refund_dispute&needs_evidence=true")

    assert response.status_code == 200
    cases = response.json()["cases"]
    assert cases
    assert {case["case_type"] for case in cases} == {"customer_refund_dispute"}
    assert all(case["evidence_status"] in {"missing", "partial"} or case["recovery_stage"] == "needs_evidence" for case in cases)


def test_recovery_actions_staff_only_gets_allowed_evidence_actions(client: TestClient, db_session: Session) -> None:
    data = seed_recovery_data(client, db_session)
    staff = create_user(client, "staff-recovery-actions@example.com", "staff")
    assign_restaurant(client, staff["id"], data["restaurant"]["id"])
    token = login(client, staff["email"])

    response = client.get("/v1/recovery/actions", headers=auth_headers(token))

    assert response.status_code == 200
    assert {action["action_type"] for action in response.json()["actions"]} == {"upload_evidence"}
