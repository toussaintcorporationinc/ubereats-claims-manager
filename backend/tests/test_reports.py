from datetime import timedelta
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.models import (
    ClaimResponseReview,
    EmailAccount,
    EmailDraft,
    EvidenceFile,
    FollowUpTask,
    InboundEmailMessage,
    UberCustomerRefundDispute,
    User,
)
from app.models.domain import utc_now


def auth_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def create_restaurant(client: TestClient, name: str = "Report Restaurant") -> dict:
    response = client.post("/v1/restaurants", json={"name": name, "sender_email": "claims@example.com"})
    assert response.status_code == 201
    return response.json()


def create_user(client: TestClient, email: str, role: str) -> dict:
    response = client.post(
        "/v1/users",
        json={
            "email": email,
            "password": "user-password",
            "full_name": f"{role.title()} Report",
            "role": role,
            "active": True,
        },
    )
    assert response.status_code == 201
    return response.json()


def assign_restaurant(client: TestClient, user_id: int, restaurant_id: int) -> None:
    response = client.post(f"/v1/users/{user_id}/restaurants", json={"restaurant_id": restaurant_id})
    assert response.status_code == 201


def login(client: TestClient, email: str, password: str = "user-password") -> str:
    response = client.post("/v1/auth/login", json={"email": email, "password": password})
    assert response.status_code == 200
    return response.json()["access_token"]


def create_order(
    client: TestClient,
    restaurant_id: int,
    order_number: str,
    amount: str,
    *,
    customer_name: str | None = "Client Test",
    status: str = "waiting_uber_response",
    result: str | None = None,
    recovered_amount: str | None = None,
    order_date: str = "2026-06-01",
) -> dict:
    response = client.post(
        "/v1/orders",
        json={
            "restaurant_id": restaurant_id,
            "uber_order_number": order_number,
            "customer_name": customer_name,
            "order_date": order_date,
            "order_amount": amount,
            "currency": "EUR",
            "accepted_by_restaurant": True,
            "prepared_before_cancellation": True,
            "status": status,
            "result": result,
            "recovered_amount": recovered_amount,
        },
    )
    assert response.status_code == 201
    return response.json()


def get_owner(db_session: Session) -> User:
    owner = db_session.scalar(select(User).where(User.email == "owner@example.com"))
    assert owner is not None
    return owner


def add_related_reporting_rows(db_session: Session, order_id: int, user_id: int) -> None:
    account = EmailAccount(
        user_id=user_id,
        provider="gmail",
        email_address="reports@example.com",
        access_token_encrypted="encrypted-access-token",
        refresh_token_encrypted="encrypted-refresh-token",
        scopes="https://www.googleapis.com/auth/gmail.readonly",
    )
    db_session.add(account)
    db_session.flush()
    db_session.add(
        EvidenceFile(
            order_id=order_id,
            evidence_type="receipt",
            original_filename="receipt.pdf",
            storage_path="private/path/receipt.pdf",
            storage_backend="local",
            mime_type="application/pdf",
            file_size=1234,
            checksum_sha256="a" * 64,
            uploaded_by_user_id=user_id,
        )
    )
    db_session.add(
        EmailDraft(
            order_id=order_id,
            draft_type="initial_claim",
            subject="Reclamation Uber Eats",
            body="Internal draft body",
            status="created",
        )
    )
    db_session.add(
        InboundEmailMessage(
            email_account_id=account.id,
            order_id=order_id,
            provider="gmail",
            provider_message_id=f"report-message-{order_id}",
            provider_thread_id=f"report-thread-{order_id}",
            from_email="support@uber.com",
            to_email="claims@example.com",
            subject="Reponse Uber",
            snippet="Traitement en cours",
            body_text="Traitement en cours",
            received_at=utc_now(),
            match_status="linked",
            match_reason="thread_id_match",
            review_status="unreviewed",
        )
    )


def add_followup(db_session: Session, order_id: int, task_type: str = "followup_1") -> FollowUpTask:
    task = FollowUpTask(
        order_id=order_id,
        task_type=task_type,
        status="pending",
        due_at=utc_now() - timedelta(hours=1),
    )
    db_session.add(task)
    db_session.commit()
    db_session.refresh(task)
    return task


def add_response_review(
    db_session: Session,
    order_id: int,
    user_id: int,
    review_type: str,
    *,
    recovered_amount: str | None = None,
) -> ClaimResponseReview:
    review = ClaimResponseReview(
        order_id=order_id,
        reviewed_by_user_id=user_id,
        review_type=review_type,
        previous_order_status="waiting_uber_response",
        new_order_status=review_type if review_type in {"accepted", "payment_confirmed", "refused"} else "manual_review",
        recovered_amount=recovered_amount,
        refusal_reason="Refus Uber" if review_type == "refused" else None,
        evidence_requested=review_type == "evidence_requested",
    )
    db_session.add(review)
    db_session.commit()
    db_session.refresh(review)
    return review


def add_customer_refund(
    db_session: Session,
    restaurant_id: int,
    *,
    order_id: str,
    amount: str,
    status: str = "needs_evidence",
) -> UberCustomerRefundDispute:
    dispute = UberCustomerRefundDispute(
        restaurant_id=restaurant_id,
        uber_store_id="store-reporting",
        uber_order_id=order_id,
        display_id=order_id,
        dispute_type="customer_refund",
        reason="refund_without_sufficient_proof",
        status=status,
        customer_refund_amount=amount,
        order_amount=amount,
        currency="EUR",
        deducted_at=utc_now().date(),
        order_date=utc_now().date(),
        evidence_required=True,
        evidence_status="missing",
        raw_payload_json={"source": "test"},
    )
    db_session.add(dispute)
    db_session.commit()
    db_session.refresh(dispute)
    return dispute


def seed_reporting_data(client: TestClient, db_session: Session) -> dict[str, dict]:
    owner = get_owner(db_session)
    first_restaurant = create_restaurant(client, "Reports A")
    second_restaurant = create_restaurant(client, "Reports B")
    pending_order = create_order(
        client,
        first_restaurant["id"],
        "UBER-REPORT-PENDING",
        "30.00",
        customer_name="Customer Hidden",
    )
    accepted_order = create_order(
        client,
        first_restaurant["id"],
        "UBER-REPORT-ACCEPTED",
        "20.00",
        status="accepted",
        result="accepted",
        recovered_amount="20.00",
    )
    refused_order = create_order(
        client,
        second_restaurant["id"],
        "UBER-REPORT-REFUSED",
        "10.00",
        status="refused",
        result="refused",
    )
    add_related_reporting_rows(db_session, pending_order["id"], owner.id)
    add_followup(db_session, pending_order["id"])
    add_response_review(db_session, accepted_order["id"], owner.id, "accepted", recovered_amount="20.00")
    add_response_review(db_session, refused_order["id"], owner.id, "refused")
    db_session.commit()
    return {
        "first_restaurant": first_restaurant,
        "second_restaurant": second_restaurant,
        "pending_order": pending_order,
        "accepted_order": accepted_order,
        "refused_order": refused_order,
    }


def test_health_public_works(unauthenticated_client: TestClient) -> None:
    response = unauthenticated_client.get("/health")

    assert response.status_code == 200
    assert response.json()["status"] == "ok"


def test_owner_can_access_commercial_summary(client: TestClient, db_session: Session) -> None:
    seed_reporting_data(client, db_session)

    response = client.get("/v1/reports/commercial-summary")

    assert response.status_code == 200
    assert response.json()["totals"]["orders_count"] == 3


def test_commercial_summary_dedupes_historical_customer_refund_duplicates(client: TestClient, db_session: Session) -> None:
    restaurant = create_restaurant(client, "Reports Refunds")
    add_customer_refund(db_session, restaurant["id"], order_id="UBER-REPORT-DUP", amount="29.99")
    add_customer_refund(db_session, restaurant["id"], order_id="UBER-REPORT-DUP", amount="29.99")

    response = client.get("/v1/reports/commercial-summary")

    assert response.status_code == 200
    customer_refunds = response.json()["customer_refunds"]
    assert customer_refunds["total_deducted_amount"] == "29.99"
    assert customer_refunds["disputes_count"] == 1


def test_commercial_summary_excludes_ignored_customer_refunds_by_default(
    client: TestClient,
    db_session: Session,
) -> None:
    restaurant = create_restaurant(client, "Reports Ignored Refunds")
    add_customer_refund(db_session, restaurant["id"], order_id="UBER-REPORT-ACTIVE", amount="12.50")
    add_customer_refund(
        db_session,
        restaurant["id"],
        order_id="UBER-REPORT-IGNORED",
        amount="99.99",
        status="ignored",
    )

    response = client.get("/v1/reports/commercial-summary")

    assert response.status_code == 200
    customer_refunds = response.json()["customer_refunds"]
    assert customer_refunds["total_deducted_amount"] == "12.50"
    assert customer_refunds["disputes_count"] == 1


def test_approved_customer_refunds_are_reported_separately_from_received(
    client: TestClient,
    db_session: Session,
) -> None:
    restaurant = create_restaurant(client, "Reports Approved Refunds")
    add_customer_refund(
        db_session,
        restaurant["id"],
        order_id="UBER-REPORT-APPROVED",
        amount="20.80",
        status="payment_to_verify",
    )
    add_customer_refund(
        db_session,
        restaurant["id"],
        order_id="UBER-REPORT-CONFIRMED",
        amount="12.50",
        status="payment_confirmed",
    )

    report_response = client.get("/v1/reports/commercial-summary")
    dashboard_response = client.get("/v1/dashboard/summary")

    assert report_response.status_code == 200
    assert dashboard_response.status_code == 200
    customer_refunds = report_response.json()["customer_refunds"]
    assert customer_refunds["total_approved_amount"] == "20.80"
    assert customer_refunds["total_recovered_amount"] == "12.50"
    assert dashboard_response.json()["total_approved_amount"] == "20.80"
    assert dashboard_response.json()["total_recovered_amount"] == "12.50"


def test_manager_summary_only_includes_assigned_restaurants(client: TestClient, db_session: Session) -> None:
    data = seed_reporting_data(client, db_session)
    manager = create_user(client, "manager-reports@example.com", "manager")
    assign_restaurant(client, manager["id"], data["first_restaurant"]["id"])
    manager_token = login(client, manager["email"])

    response = client.get("/v1/reports/commercial-summary", headers=auth_headers(manager_token))

    assert response.status_code == 200
    summary = response.json()
    assert summary["totals"]["orders_count"] == 2
    assert {row["restaurant_name"] for row in summary["by_restaurant"]} == {"Reports A"}


def test_staff_cannot_access_reports_or_exports(client: TestClient) -> None:
    staff = create_user(client, "staff-reports@example.com", "staff")
    staff_token = login(client, staff["email"])

    report_response = client.get("/v1/reports/commercial-summary", headers=auth_headers(staff_token))
    export_response = client.get("/v1/reports/export/orders.csv", headers=auth_headers(staff_token))

    assert report_response.status_code == 403
    assert export_response.status_code == 403


def test_unauthorized_restaurant_filter_is_refused(client: TestClient, db_session: Session) -> None:
    data = seed_reporting_data(client, db_session)
    manager = create_user(client, "manager-reports-filter@example.com", "manager")
    assign_restaurant(client, manager["id"], data["first_restaurant"]["id"])
    manager_token = login(client, manager["email"])

    response = client.get(
        f"/v1/reports/commercial-summary?restaurant_id={data['second_restaurant']['id']}",
        headers=auth_headers(manager_token),
    )

    assert response.status_code == 403


def test_summary_calculates_commercial_totals_and_breakdowns(client: TestClient, db_session: Session) -> None:
    seed_reporting_data(client, db_session)

    response = client.get("/v1/reports/commercial-summary")

    assert response.status_code == 200
    summary = response.json()
    assert summary["totals"]["total_claimed_amount"] == "60.00"
    assert summary["totals"]["total_recovered_amount"] == "20.00"
    assert summary["totals"]["total_pending_amount"] == "30.00"
    assert summary["totals"]["total_refused_amount"] == "10.00"
    assert summary["totals"]["success_rate"] == "0.50"
    assert {row["key"] for row in summary["by_status"]} >= {"accepted", "refused", "waiting_uber_response"}
    assert {row["key"] for row in summary["by_result"]} >= {"accepted", "refused", "none"}
    assert len(summary["by_restaurant"]) == 2


def test_reports_orders_returns_counts_without_customer_name_by_default(client: TestClient, db_session: Session) -> None:
    data = seed_reporting_data(client, db_session)

    response = client.get(f"/v1/reports/orders?restaurant_id={data['first_restaurant']['id']}")

    assert response.status_code == 200
    orders = response.json()["orders"]
    pending = next(order for order in orders if order["uber_order_number"] == "UBER-REPORT-PENDING")
    assert "customer_name" not in pending
    assert pending["evidence_count"] == 1
    assert pending["drafts_count"] == 1
    assert pending["inbound_messages_count"] == 1
    assert pending["response_reviews_count"] == 0


def test_reports_orders_can_include_customer_names_for_owner_manager(client: TestClient, db_session: Session) -> None:
    seed_reporting_data(client, db_session)

    response = client.get("/v1/reports/orders?include_customer_names=true")

    assert response.status_code == 200
    pending = next(order for order in response.json()["orders"] if order["uber_order_number"] == "UBER-REPORT-PENDING")
    assert pending["customer_name"] == "Customer Hidden"


def test_export_orders_csv_works_and_omits_sensitive_fields(client: TestClient, db_session: Session) -> None:
    seed_reporting_data(client, db_session)

    response = client.get("/v1/reports/export/orders.csv")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    content = response.text
    assert "UBER-REPORT-PENDING" in content
    assert "customer_name" not in content
    assert "storage_path" not in content
    assert "access_token" not in content
    assert "refresh_token" not in content


def test_export_orders_xlsx_works(client: TestClient, db_session: Session) -> None:
    seed_reporting_data(client, db_session)

    response = client.get("/v1/reports/export/orders.xlsx")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert workbook.sheetnames == ["Orders"]
    assert workbook["Orders"]["A1"].value == "order_id"


def test_export_commercial_summary_xlsx_contains_multiple_sheets(client: TestClient, db_session: Session) -> None:
    seed_reporting_data(client, db_session)

    response = client.get("/v1/reports/export/commercial-summary.xlsx")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert {"Summary", "By Restaurant", "By Status", "By Result", "Followups", "Responses"} <= set(workbook.sheetnames)


def test_export_respects_max_rows(client: TestClient, db_session: Session, monkeypatch) -> None:
    seed_reporting_data(client, db_session)
    monkeypatch.setenv("EXPORT_MAX_ROWS", "1")
    get_settings.cache_clear()

    response = client.get("/v1/reports/export/orders.csv")

    assert response.status_code == 413
    get_settings.cache_clear()


def test_reports_followups_works(client: TestClient, db_session: Session) -> None:
    seed_reporting_data(client, db_session)

    response = client.get("/v1/reports/followups")

    assert response.status_code == 200
    followups = response.json()["followups"]
    assert len(followups) == 1
    assert followups[0]["task_type"] == "followup_1"


def test_reports_responses_works(client: TestClient, db_session: Session) -> None:
    seed_reporting_data(client, db_session)

    response = client.get("/v1/reports/responses")

    assert response.status_code == 200
    assert {row["review_type"] for row in response.json()["responses"]} == {"accepted", "refused"}


def test_response_and_followup_exports_work(client: TestClient, db_session: Session) -> None:
    seed_reporting_data(client, db_session)

    followups_response = client.get("/v1/reports/export/followups.csv")
    responses_response = client.get("/v1/reports/export/responses.csv")

    assert followups_response.status_code == 200
    assert "followup_1" in followups_response.text
    assert responses_response.status_code == 200
    assert "accepted" in responses_response.text


def test_dashboard_summary_includes_reporting_additions(client: TestClient, db_session: Session) -> None:
    seed_reporting_data(client, db_session)

    response = client.get("/v1/dashboard/summary")

    assert response.status_code == 200
    data = response.json()
    assert data["success_rate"] == "0.50"
    assert data["top_restaurants_by_claimed_amount"]
    assert data["top_restaurants_by_pending_amount"]


def test_date_and_amount_filters_apply_to_report_orders(client: TestClient, db_session: Session) -> None:
    seed_reporting_data(client, db_session)

    response = client.get("/v1/reports/orders?date_from=2026-06-01&date_to=2026-06-01&min_amount=25&max_amount=35")

    assert response.status_code == 200
    orders = response.json()["orders"]
    assert [order["uber_order_number"] for order in orders] == ["UBER-REPORT-PENDING"]
