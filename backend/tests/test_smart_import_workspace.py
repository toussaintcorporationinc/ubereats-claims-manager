from datetime import datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from zipfile import ZipFile

from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import (
    AuditLog,
    ClaimOrder,
    EvidenceFile,
    EvidenceImportBatch,
    EvidenceImportedFile,
    EvidenceRequestTask,
    SmartImportPreviewBatch,
    UberCustomerRefundDispute,
    UberFinancialTransaction,
    UberOrderSnapshot,
    UberReportingImportBatch,
    UberReportingImportRow,
    UberStoreMapping,
    Restaurant,
    User,
)
from app.models.domain import utc_now
import app.services.evidence_ai_analysis_service as evidence_analysis_service


def test_health_public_works(unauthenticated_client: TestClient) -> None:
    response = unauthenticated_client.get("/health")
    assert response.status_code == 200


def test_smart_import_detects_uber_report_without_filename(client: TestClient) -> None:
    csv_content = (
        "Store id,Nom du restaurant,Id. de la commande,Date de la commande,Statut de la commande,Ventes (TVA incluse),Devise\n"
        "store-1,Restaurant Test TENNET,UBER-SMART-001,2026-05-01,canceled,24.90,EUR\n"
    )
    response = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("download.csv", csv_content.encode("utf-8"), "text/csv"))],
    )

    assert response.status_code == 201
    file_preview = response.json()["files"][0]
    assert file_preview["detected_category"] == "uber_reporting"
    assert file_preview["recommended_action"] == "import_uber_reporting"


def test_smart_import_detects_two_line_uber_header_and_restaurant_period(client: TestClient) -> None:
    csv_content = (
        "Descriptions longues Uber Eats Manager,,,,,,\n"
        "Id. du restaurant,Nom du restaurant,Id. de la commande,Date de la commande,Statut de la commande,Ventes (TVA incluse),Devise\n"
        "store-2,Restaurant Test Header,UBER-SMART-002,01/05/2026,annule,31,EUR\n"
        "store-2,Restaurant Test Header,UBER-SMART-003,03/05/2026,canceled,18.50,EUR\n"
    )
    response = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("897e7ee9-f500.csv", csv_content.encode("utf-8"), "text/csv"))],
    )

    assert response.status_code == 201
    file_preview = response.json()["files"][0]
    assert file_preview["header_row_number"] == 2
    assert file_preview["skipped_preamble_rows"] == 1
    assert file_preview["detected_restaurant_name"] == "Restaurant Test Header"
    assert file_preview["detected_date_from"] == "2026-05-01"
    assert file_preview["detected_date_to"] == "2026-05-03"


def test_smart_import_detects_xlsx_two_line_header(client: TestClient) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Descriptions longues Uber"])
    sheet.append(["Id. du restaurant", "Nom du restaurant", "Id. de la commande", "Date de la commande", "Montant total"])
    sheet.append(["store-3", "Restaurant XLSX", "UBER-XLSX-001", "2026-05-10", "12,50"])
    buffer = BytesIO()
    workbook.save(buffer)

    response = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("export.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
    )

    assert response.status_code == 201
    file_preview = response.json()["files"][0]
    assert file_preview["detected_category"] == "uber_reporting"
    assert file_preview["header_row_number"] == 2


def test_smart_import_detects_zip_and_evidence_image(client: TestClient) -> None:
    response = client.post(
        "/v1/smart-import/preview",
        files=[
            ("files", ("preuve.zip", b"PK\x03\x04", "application/zip")),
            ("files", ("IMG_1234.jpg", b"fake-image", "image/jpeg")),
        ],
    )

    assert response.status_code == 201
    files = response.json()["files"]
    assert files[0]["detected_category"] == "zip"
    assert files[0]["recommended_action"] == "import_evidence_bulk"
    assert files[1]["detected_category"] == "evidence"


def test_smart_import_unknown_becomes_manual_review(client: TestClient) -> None:
    response = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("export.csv", b"foo,bar\nbaz,qux\n", "text/csv"))],
    )

    assert response.status_code == 201
    file_preview = response.json()["files"][0]
    assert file_preview["detected_category"] == "unknown"
    assert file_preview["recommended_action"] == "manual_review"


def test_smart_import_detects_official_uber_accuracy_orders_as_adjustments(client: TestClient) -> None:
    csv_content = official_inaccurate_orders_csv()

    response = client.post(
        "/v1/smart-import/preview",
        files=[
            (
                "files",
                (
                    "download.csv",
                    csv_content.encode("utf-8"),
                    "text/csv",
                ),
            )
        ],
    )

    assert response.status_code == 201
    file_preview = response.json()["files"][0]
    assert file_preview["detected_category"] == "uber_reporting"
    assert file_preview["detected_report_type"] == "adjustments_report"
    assert file_preview["recommended_action"] == "import_uber_reporting"
    assert "Remboursement pris en charge par le commerçant" in file_preview["detected_columns"]


def test_smart_confirm_official_uber_accuracy_orders_creates_refund_transaction(
    client: TestClient,
    db_session: Session,
) -> None:
    restaurant = Restaurant(name="Krousty Bat", sender_email="tiramisumaisonfrance@example.com")
    db_session.add(restaurant)
    db_session.flush()
    db_session.add(
        UberStoreMapping(
            restaurant_id=restaurant.id,
            uber_store_id="store-accuracy",
            uber_store_name="Krousty Bat",
            active=True,
        )
    )
    db_session.commit()
    csv_content = official_inaccurate_orders_csv()
    preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("inaccurate_orders_v3_2026-01-01_2026-01-31.csv", csv_content.encode("utf-8"), "text/csv"))],
    ).json()

    response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [
                {
                    "file_id": preview["files"][0]["id"],
                    "action": "import_uber_reporting",
                    "report_type": "adjustments_report",
                }
            ],
        },
    )

    assert response.status_code == 200
    routed = response.json()["routed_files"][0]
    assert routed["destination_type"] == "uber_reporting_batch"
    assert routed["created_transactions_count"] == 1
    assert routed["skipped_rows"] == 0
    transaction = db_session.scalar(select(UberFinancialTransaction))
    assert transaction is not None
    assert transaction.restaurant_id == restaurant.id
    assert transaction.uber_order_id == "PROCESS-ACCURACY-001"
    assert transaction.transaction_type == "customer_refund"
    assert str(transaction.amount) == "-12.50"
    assert transaction.raw_payload_json["raw_data"]["probleme avec la commande"] == "Article manquant"

    detect_response = client.post("/v1/customer-refunds/detect", json={"restaurant_id": restaurant.id})
    assert detect_response.status_code == 200
    dispute = db_session.scalar(select(UberCustomerRefundDispute))
    assert dispute is not None
    assert dispute.dispute_type == "missing_item"
    assert str(dispute.customer_refund_amount) == "12.50"


def test_smart_import_routes_top_inaccurate_items_as_uber_reporting_source(
    client: TestClient,
    db_session: Session,
) -> None:
    csv_content = (
        "Restaurant,Id. externe du restaurant,Pays,Code pays,Ville,Articles incorrects,Personnalisations incorrectes,"
        "Id. externe de l'article,Données externes,Problème avec la commande,Problème avec le plat,Nombre\n"
        "Krousty Bat,store-accuracy,France,FR,Paris,Burger Test,,item-1,sku-1,Article manquant,,4\n"
    )
    preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("top_inaccurate_items_v3_2026-01-01_2026-01-31.csv", csv_content.encode("utf-8"), "text/csv"))],
    ).json()
    file_preview = preview["files"][0]
    assert file_preview["detected_category"] == "uber_reporting"
    assert file_preview["detected_report_type"] == "adjustments_report"
    assert file_preview["recommended_action"] == "import_uber_reporting"
    assert "missing_order_level_columns" in file_preview["warnings"]

    response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [{"file_id": file_preview["id"], "action": "import_uber_reporting", "report_type": "adjustments_report"}],
        },
    )

    assert response.status_code == 200
    routed = response.json()["routed_files"][0]
    assert routed["destination_type"] == "uber_reporting_batch"
    assert routed["destination_url"] == f"/uber/reporting/{routed['destination_id']}"
    assert routed["created_transactions_count"] == 0
    assert routed["skipped_rows"] == 1
    batch = db_session.get(UberReportingImportBatch, routed["destination_id"])
    assert batch is not None
    assert batch.report_type == "adjustments_report"
    assert batch.invalid_rows == 1


def test_smart_import_routes_order_accuracy_workbook_as_uber_reporting_source(
    client: TestClient,
    db_session: Session,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    sheet.append(["Période d'évaluation"])
    sheet.append(["2026-01-01 - 2026-01-31"])
    sheet.append(["Commandes incorrectes", 3])
    buffer = BytesIO()
    workbook.save(buffer)
    preview = client.post(
        "/v1/smart-import/preview",
        files=[
            (
                "files",
                (
                    "order_accuracy_analytics_2026-01-01_2026-01-31.xlsx",
                    buffer.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            )
        ],
    ).json()
    file_preview = preview["files"][0]
    assert file_preview["detected_category"] == "uber_reporting"
    assert file_preview["detected_report_type"] == "adjustments_report"
    assert file_preview["recommended_action"] == "import_uber_reporting"

    response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [{"file_id": file_preview["id"], "action": "import_uber_reporting", "report_type": "adjustments_report"}],
        },
    )

    assert response.status_code == 200
    routed = response.json()["routed_files"][0]
    assert routed["destination_type"] == "uber_reporting_batch"
    assert routed["destination_url"] == f"/uber/reporting/{routed['destination_id']}"
    assert routed["created_transactions_count"] == 0
    batch = db_session.get(UberReportingImportBatch, routed["destination_id"])
    assert batch is not None
    assert batch.report_type == "adjustments_report"


def test_smart_confirm_order_accuracy_positive_refund_amount_becomes_negative_transaction(
    client: TestClient,
    db_session: Session,
) -> None:
    restaurant = Restaurant(name="Krousty Bat", sender_email="tiramisumaisonfrance@example.com")
    db_session.add(restaurant)
    db_session.flush()
    db_session.add(
        UberStoreMapping(
            restaurant_id=restaurant.id,
            uber_store_id="store-accuracy-positive",
            uber_store_name="Krousty Bat",
            active=True,
        )
    )
    db_session.commit()
    csv_content = (
        "Store UUID,Store Name,Order UUID,Order Date,Issue,Refund Amount,Currency\n"
        "store-accuracy-positive,Krousty Bat,ORDER-ACCURACY-POSITIVE-001,2026-01-08,Missing item,12.50,EUR\n"
    )
    preview = client.post(
        "/v1/smart-import/preview",
        files=[
            (
                "files",
                (
                    "order_accuracy_analytics_2026-01-01_2026-01-31.csv",
                    csv_content.encode("utf-8"),
                    "text/csv",
                ),
            )
        ],
    ).json()
    file_preview = preview["files"][0]
    assert file_preview["detected_category"] == "uber_reporting"
    assert file_preview["recommended_action"] == "import_uber_reporting"

    response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [
                {
                    "file_id": file_preview["id"],
                    "action": "import_uber_reporting",
                    "report_type": "adjustments_report",
                }
            ],
        },
    )

    assert response.status_code == 200
    routed = response.json()["routed_files"][0]
    assert routed["created_transactions_count"] == 1
    transaction = db_session.scalar(
        select(UberFinancialTransaction).where(UberFinancialTransaction.uber_order_id == "ORDER-ACCURACY-POSITIVE-001")
    )
    assert transaction is not None
    assert transaction.transaction_type == "customer_refund"
    assert transaction.amount == Decimal("-12.50")


def test_smart_confirm_routes_uber_report_to_reporting_batch(client: TestClient, db_session: Session) -> None:
    restaurant = Restaurant(name="Restaurant Route", sender_email="route@example.com")
    db_session.add(restaurant)
    db_session.flush()
    db_session.add(
        UberStoreMapping(
            restaurant_id=restaurant.id,
            uber_store_id="store-route",
            uber_store_name="Restaurant Route",
            active=True,
        )
    )
    db_session.commit()
    csv_content = (
        "Descriptions longues Uber Eats Manager,,,,,,\n"
        "Id. du restaurant,Nom du restaurant,Id. de la commande,Date de la commande,Statut de la commande,Ventes (TVA incluse),Devise\n"
        "store-route,Restaurant Route,UBER-ROUTE-001,01/05/2026,canceled,31,EUR\n"
    )
    preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("download.csv", csv_content.encode("utf-8"), "text/csv"))],
    ).json()

    response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [{"file_id": preview["files"][0]["id"], "action": "import_uber_reporting", "report_type": "orders_report"}],
        },
    )

    assert response.status_code == 200
    routed = response.json()["routed_files"][0]
    assert routed["destination_type"] == "uber_reporting_batch"
    assert routed["destination_url"] == f"/uber/reporting/{routed['destination_id']}"
    assert routed["processing_status"] == "confirmed"
    assert routed["created_snapshots_count"] == 1
    assert routed["skipped_rows"] == 0
    batch = db_session.get(UberReportingImportBatch, routed["destination_id"])
    assert batch is not None
    assert batch.status == "confirmed"
    assert batch.report_type == "orders_report"
    rows = db_session.scalars(select(UberReportingImportRow).where(UberReportingImportRow.batch_id == batch.id)).all()
    assert rows
    assert all(row.status == "created" for row in rows)


def test_smart_confirm_forced_restaurant_applies_unmapped_uber_report(client: TestClient, db_session: Session) -> None:
    restaurant = Restaurant(name="Restaurant Force Smart", sender_email="force@example.com")
    db_session.add(restaurant)
    db_session.commit()
    db_session.refresh(restaurant)
    csv_content = (
        "Id. du restaurant,Nom du restaurant,Id. de la commande,Date de la commande,Statut de la commande,Ventes (TVA incluse),Devise\n"
        "store-force-smart,Restaurant Force Smart,UBER-FORCE-001,01/05/2026,canceled,22.50,EUR\n"
    )
    preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("force.csv", csv_content.encode("utf-8"), "text/csv"))],
    ).json()

    response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [
                {
                    "file_id": preview["files"][0]["id"],
                    "action": "import_uber_reporting",
                    "report_type": "orders_report",
                    "restaurant_id": restaurant.id,
                }
            ],
        },
    )

    assert response.status_code == 200
    routed = response.json()["routed_files"][0]
    assert routed["processing_status"] == "confirmed"
    assert routed["created_snapshots_count"] == 1
    batch = db_session.get(UberReportingImportBatch, routed["destination_id"])
    assert batch is not None
    rows = db_session.scalars(select(UberReportingImportRow).where(UberReportingImportRow.batch_id == batch.id)).all()
    assert rows[0].normalized_data["restaurant_id"] == restaurant.id
    assert rows[0].status == "created"


def test_smart_confirm_preserves_per_row_store_mapping_when_restaurant_override_supplied(
    client: TestClient,
    db_session: Session,
) -> None:
    krousty = Restaurant(name="Krousty Bat", sender_email="krousty@example.com")
    asian = Restaurant(name="Asian Passion", sender_email="asian@example.com")
    db_session.add_all([krousty, asian])
    db_session.flush()
    db_session.add_all(
        [
            UberStoreMapping(
                restaurant_id=krousty.id,
                uber_store_id="store-krousty",
                uber_store_name="Krousty Bat",
                active=True,
            ),
            UberStoreMapping(
                restaurant_id=asian.id,
                uber_store_id="store-asian",
                uber_store_name="Asian Passion",
                active=True,
            ),
        ]
    )
    db_session.commit()
    csv_content = (
        "Id. du restaurant,Nom du restaurant,Id. de la commande,Date de la commande,Statut de la commande,Ventes (TVA incluse),Devise\n"
        "store-krousty,Krousty Bat,UBER-MULTI-KROUSTY-001,01/05/2026,canceled,22.50,EUR\n"
        "store-asian,Asian Passion,UBER-MULTI-ASIAN-001,01/05/2026,canceled,18.90,EUR\n"
    )
    preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("multi-restaurants.csv", csv_content.encode("utf-8"), "text/csv"))],
    ).json()

    response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [
                {
                    "file_id": preview["files"][0]["id"],
                    "action": "import_uber_reporting",
                    "report_type": "orders_report",
                    "restaurant_id": krousty.id,
                }
            ],
        },
    )

    assert response.status_code == 200
    routed = response.json()["routed_files"][0]
    assert routed["created_snapshots_count"] == 2
    asian_snapshot = db_session.scalar(
        select(UberOrderSnapshot).where(UberOrderSnapshot.uber_order_id == "UBER-MULTI-ASIAN-001")
    )
    assert asian_snapshot is not None
    assert asian_snapshot.restaurant_id == asian.id
    krousty_snapshot = db_session.scalar(
        select(UberOrderSnapshot).where(UberOrderSnapshot.uber_order_id == "UBER-MULTI-KROUSTY-001")
    )
    assert krousty_snapshot is not None
    assert krousty_snapshot.restaurant_id == krousty.id


def test_smart_confirm_matches_restaurant_by_exact_store_name_when_store_id_missing(
    client: TestClient,
    db_session: Session,
) -> None:
    krousty = Restaurant(name="Krousty Bat", sender_email="krousty@example.com")
    asian = Restaurant(name="Asian Passion", sender_email="asian@example.com")
    db_session.add_all([krousty, asian])
    db_session.commit()
    csv_content = (
        "Nom du restaurant,Id. de la commande,Date de la commande,Statut de la commande,Ventes (TVA incluse),Devise\n"
        "Krousty Bat,UBER-NAME-KROUSTY-001,01/05/2026,canceled,22.50,EUR\n"
        "Asian Passion,UBER-NAME-ASIAN-001,01/05/2026,canceled,18.90,EUR\n"
    )
    preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("restaurants-without-store-id.csv", csv_content.encode("utf-8"), "text/csv"))],
    ).json()

    response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [{"file_id": preview["files"][0]["id"], "action": "import_uber_reporting", "report_type": "orders_report"}],
        },
    )

    assert response.status_code == 200
    routed = response.json()["routed_files"][0]
    assert routed["created_snapshots_count"] == 2
    asian_snapshot = db_session.scalar(
        select(UberOrderSnapshot).where(UberOrderSnapshot.uber_order_id == "UBER-NAME-ASIAN-001")
    )
    assert asian_snapshot is not None
    assert asian_snapshot.restaurant_id == asian.id
    assert asian_snapshot.uber_store_id == f"restaurant-name:{asian.id}"


def test_smart_import_preview_marks_exact_duplicate_by_checksum(client: TestClient, db_session: Session) -> None:
    csv_content = (
        "Id. du restaurant,Nom du restaurant,Id. de la commande,Date de la commande,Statut de la commande,Ventes (TVA incluse),Devise\n"
        "store-dup,Restaurant Dup,UBER-DUP-001,01/05/2026,canceled,31,EUR\n"
    )

    response = client.post(
        "/v1/smart-import/preview",
        files=[
            ("files", ("download (1).csv", csv_content.encode("utf-8"), "text/csv")),
            ("files", ("download.csv", csv_content.encode("utf-8"), "text/csv")),
        ],
    )

    assert response.status_code == 201
    files = sorted(response.json()["files"], key=lambda item: item["original_filename"])
    canonical = next(item for item in files if item["original_filename"] == "download.csv")
    duplicate = next(item for item in files if item["original_filename"] == "download (1).csv")
    assert canonical["recommended_action"] == "import_uber_reporting"
    assert duplicate["recommended_action"] == "ignore"
    assert duplicate["status"] == "ignored"
    assert duplicate["destination_type"] == "duplicate_ignored"
    assert duplicate["destination_id"] == canonical["id"]
    assert "exact_duplicate_ignored" in duplicate["warnings"]

    duplicate_audit = db_session.scalar(
        select(AuditLog).where(
            AuditLog.entity_type == "smart_import_preview_file",
            AuditLog.action == "smart_import_file.duplicate_ignored",
            AuditLog.entity_id == duplicate["id"],
        )
    )
    assert duplicate_audit is not None


def test_smart_import_preview_ignores_file_already_seen_in_previous_batch(client: TestClient) -> None:
    csv_content = (
        "Id. du restaurant,Nom du restaurant,Id. de la commande,Date de la commande,Statut de la commande,Ventes (TVA incluse),Devise\n"
        "store-dup-history,Restaurant Dup History,UBER-DUP-HISTORY-001,01/05/2026,canceled,31,EUR\n"
    )
    first_preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("download.csv", csv_content.encode("utf-8"), "text/csv"))],
    ).json()
    first_file = first_preview["files"][0]
    confirm_response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": first_preview["batch_preview_id"],
            "files": [{"file_id": first_file["id"], "action": "import_uber_reporting", "report_type": "orders_report"}],
        },
    )
    assert confirm_response.status_code == 200

    second_response = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("download-again.csv", csv_content.encode("utf-8"), "text/csv"))],
    )

    assert second_response.status_code == 201
    duplicate = second_response.json()["files"][0]
    assert duplicate["recommended_action"] == "ignore"
    assert duplicate["status"] == "ignored"
    assert duplicate["destination_type"] == "duplicate_ignored"
    assert duplicate["destination_id"] == first_file["id"]
    assert "exact_duplicate_ignored" in duplicate["warnings"]


def test_smart_import_preview_does_not_ignore_previous_file_that_was_never_routed(client: TestClient) -> None:
    csv_content = (
        "Id. du restaurant,Nom du restaurant,Id. de la commande,Date de la commande,Statut de la commande,Ventes (TVA incluse),Devise\n"
        "store-dup-stuck,Restaurant Dup Stuck,UBER-DUP-STUCK-001,01/05/2026,canceled,31,EUR\n"
    )
    first_preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("download.csv", csv_content.encode("utf-8"), "text/csv"))],
    )
    assert first_preview.status_code == 201

    second_response = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("download-again.csv", csv_content.encode("utf-8"), "text/csv"))],
    )

    assert second_response.status_code == 201
    duplicate = second_response.json()["files"][0]
    assert duplicate["recommended_action"] == "import_uber_reporting"
    assert duplicate["status"] == "previewed"
    assert duplicate["destination_type"] is None


def test_smart_confirm_does_not_route_exact_duplicate_even_if_forced(client: TestClient, db_session: Session) -> None:
    csv_content = (
        "Id. du restaurant,Nom du restaurant,Id. de la commande,Date de la commande,Statut de la commande,Ventes (TVA incluse),Devise\n"
        "store-dup-force,Restaurant Dup Force,UBER-DUP-FORCE-001,01/05/2026,canceled,31,EUR\n"
    )
    preview = client.post(
        "/v1/smart-import/preview",
        files=[
            ("files", ("copy.csv", csv_content.encode("utf-8"), "text/csv")),
            ("files", ("copy-again.csv", csv_content.encode("utf-8"), "text/csv")),
        ],
    ).json()

    response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [
                {"file_id": preview["files"][0]["id"], "action": "import_uber_reporting", "report_type": "orders_report"},
                {"file_id": preview["files"][1]["id"], "action": "import_uber_reporting", "report_type": "orders_report"},
            ],
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert len(payload["routed_files"]) == 1
    assert len(payload["ignored_files"]) == 1
    assert payload["ignored_files"][0]["destination_type"] == "duplicate_ignored"
    reporting_batches = db_session.scalars(select(UberReportingImportBatch)).all()
    assert len(reporting_batches) == 1
    batch = db_session.get(SmartImportPreviewBatch, preview["batch_preview_id"])
    assert batch is not None
    assert len([file for file in batch.files if file.status == "routed"]) == 1
    assert len([file for file in batch.files if file.status == "ignored"]) == 1


def test_smart_confirm_routes_evidence_file_to_evidence_import(client: TestClient, db_session: Session) -> None:
    preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("ticket-test.jpg", b"fake image bytes", "image/jpeg"))],
    ).json()

    response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [{"file_id": preview["files"][0]["id"], "action": "import_evidence_bulk"}],
        },
    )

    assert response.status_code == 200
    routed = response.json()["routed_files"][0]
    assert routed["destination_type"] == "evidence_import_batch"
    assert routed["destination_url"] == f"/evidence-imports/{routed['destination_id']}"
    assert routed["processing_status"] == "analyzed"
    assert routed["analyzed_files_count"] == 1
    batch = db_session.get(EvidenceImportBatch, routed["destination_id"])
    assert batch is not None
    assert batch.status == "analyzed"
    assert batch.stored_files_count == 1


def test_smart_confirm_auto_attaches_exact_evidence_task(client: TestClient, db_session: Session) -> None:
    restaurant, order = create_restaurant_order_and_task(db_session)
    preview = client.post(
        "/v1/smart-import/preview",
        files=[
            (
                "files",
                (
                    "ticket-krousty-UBER-NEXT-001.pdf",
                    b"receipt UBER-NEXT-001 Client Test 125.00",
                    "application/pdf",
                ),
            )
        ],
    ).json()

    response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [
                {
                    "file_id": preview["files"][0]["id"],
                    "action": "import_evidence_bulk",
                    "restaurant_id": restaurant.id,
                }
            ],
        },
    )

    assert response.status_code == 200
    routed = response.json()["routed_files"][0]
    assert routed["destination_type"] == "evidence_import_batch"
    assert routed["auto_matched_count"] == 1
    task = db_session.scalar(select(EvidenceRequestTask).where(EvidenceRequestTask.order_id == order.id))
    assert task is not None
    db_session.refresh(task)
    assert task.status == "completed"
    assert db_session.scalar(select(EvidenceImportBatch).where(EvidenceImportBatch.id == routed["destination_id"])) is not None


def test_smart_confirm_duplicate_evidence_creates_visible_ignored_file(client: TestClient, db_session: Session) -> None:
    first = client.post(
        "/v1/evidence-imports",
        files=[("files", ("ticket-original.jpg", b"same smart evidence bytes", "image/jpeg"))],
    )
    assert first.status_code == 201

    preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("ticket-again.jpg", b"same smart evidence bytes", "image/jpeg"))],
    ).json()

    response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [{"file_id": preview["files"][0]["id"], "action": "import_evidence_bulk"}],
        },
    )

    assert response.status_code == 200
    routed = response.json()["routed_files"][0]
    batch = db_session.get(EvidenceImportBatch, routed["destination_id"])
    assert batch is not None
    assert batch.status == "analyzed"
    assert batch.stored_files_count == 0
    assert batch.duplicate_files_count == 1
    imported_files = db_session.scalars(select(EvidenceImportedFile).where(EvidenceImportedFile.batch_id == batch.id)).all()
    assert len(imported_files) == 1
    assert imported_files[0].status == "ignored"
    assert imported_files[0].original_filename == "ticket-again.jpg"


def test_smart_confirm_routes_zip_to_evidence_import(client: TestClient, db_session: Session) -> None:
    archive = BytesIO()
    with ZipFile(archive, "w") as zip_file:
        zip_file.writestr("ticket-a.jpg", b"fake image")
    preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("preuve.zip", archive.getvalue(), "application/zip"))],
    ).json()

    response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [{"file_id": preview["files"][0]["id"], "action": "import_evidence_bulk"}],
        },
    )

    assert response.status_code == 200
    routed = response.json()["routed_files"][0]
    batch = db_session.get(EvidenceImportBatch, routed["destination_id"])
    assert batch is not None
    assert batch.source_type == "zip_upload"
    assert batch.stored_files_count == 1
    assert batch.status == "analyzed"


def test_smart_confirm_manual_review_and_ignore_keep_audit_state(client: TestClient, db_session: Session) -> None:
    preview = client.post(
        "/v1/smart-import/preview",
        files=[
            ("files", ("unknown.csv", b"foo,bar\nbaz,qux\n", "text/csv")),
            ("files", ("ignore.jpg", b"fake image bytes", "image/jpeg")),
        ],
    ).json()

    response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [
                {"file_id": preview["files"][0]["id"], "action": "manual_review"},
                {"file_id": preview["files"][1]["id"], "action": "ignore"},
            ],
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["manual_review_files"][0]["destination_type"] == "manual_review"
    assert payload["ignored_files"][0]["destination_type"] == "ignored"
    batch = db_session.get(SmartImportPreviewBatch, preview["batch_preview_id"])
    assert batch is not None
    assert {file.status for file in batch.files} == {"manual_review", "ignored"}


def test_workspace_machine_recovers_manual_review_smart_import_file_into_proof_pipeline(
    client: TestClient,
    db_session: Session,
) -> None:
    restaurant = Restaurant(name="Krousty Bat", sender_email="tiramisumaisonfrance@example.com")
    db_session.add(restaurant)
    db_session.commit()
    proof_text = (
        "Restaurant: Krousty Bat\n"
        "Client: Client Repris\n"
        "Commande: GO-RECOVER-123\n"
        "Date: 15/06/2026\n"
        "Montant total: 27,40 EUR\n"
        "Remboursement client - article manquant\n"
    )
    preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("source-a-verifier.csv", proof_text.encode("utf-8"), "text/csv"))],
    ).json()
    confirm_response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [{"file_id": preview["files"][0]["id"], "action": "manual_review"}],
        },
    )
    assert confirm_response.status_code == 200
    assert confirm_response.json()["manual_review_files"][0]["destination_type"] == "manual_review"

    response = client.post(
        "/v1/workspace/machine/run",
        json={
            "trigger": "refunds",
            "smart_import_batch_id": preview["batch_preview_id"],
            "sync_gmail": False,
            "run_autopilot": False,
        },
    )

    assert response.status_code == 200
    stages = {stage["name"]: stage for stage in response.json()["stages"]}
    assert stages["smart_import_recovery"]["created_count"] == 1
    assert stages["evidence"]["processed_count"] >= 1
    assert stages["proof_intake"]["created_count"] == 1
    batch = db_session.get(SmartImportPreviewBatch, preview["batch_preview_id"])
    assert batch is not None
    preview_file = batch.files[0]
    assert preview_file.status == "routed"
    assert preview_file.destination_type == "evidence_import_batch"
    evidence_batch = db_session.get(EvidenceImportBatch, preview_file.destination_id)
    assert evidence_batch is not None
    assert evidence_batch.status == "analyzed"
    order = db_session.scalar(select(ClaimOrder).where(ClaimOrder.uber_order_number == "GO-RECOVER-123"))
    assert order is not None
    assert order.restaurant_id == restaurant.id
    assert order.customer_name == "Client Repris"
    assert order.order_date.isoformat() == "2026-06-15"
    assert str(order.order_amount) == "27.40"


def test_smart_confirm_expired_preview_refused(client: TestClient, db_session: Session) -> None:
    preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("ticket-test.jpg", b"fake image bytes", "image/jpeg"))],
    ).json()
    batch = db_session.get(SmartImportPreviewBatch, preview["batch_preview_id"])
    assert batch is not None
    batch.expires_at = utc_now() - timedelta(hours=1)
    db_session.commit()

    response = client.post(
        "/v1/smart-import/confirm",
        json={"batch_preview_id": preview["batch_preview_id"]},
    )

    assert response.status_code == 410


def test_workspace_next_actions_returns_priorities_for_owner(client: TestClient, db_session: Session) -> None:
    restaurant, order = create_restaurant_order_and_task(db_session)

    response = client.get("/v1/workspace/next-actions")

    assert response.status_code == 200
    payload = response.json()
    urgent_titles = [item["title"] for item in payload["urgent"]]
    all_urls = [item["action_url"] for bucket in payload.values() for item in bucket]
    assert any("Ticket" in title or "preuve" in title.lower() for title in urgent_titles)
    assert "/smart-import" in all_urls
    assert "/recovery" in all_urls
    assert "/reports" in all_urls
    assert "/autopilot" in all_urls
    assert restaurant.id == order.restaurant_id


def test_staff_next_actions_limited_to_evidence(client: TestClient, db_session: Session) -> None:
    restaurant, _order = create_restaurant_order_and_task(db_session)
    staff = client.post(
        "/v1/users",
        json={
            "email": "staff-next-actions@example.com",
            "password": "staff-password",
            "full_name": "Staff Next Actions",
            "role": "staff",
            "active": True,
        },
    ).json()
    assign_response = client.post(f"/v1/users/{staff['id']}/restaurants", json={"restaurant_id": restaurant.id})
    assert assign_response.status_code == 201
    login_response = client.post("/v1/auth/login", json={"email": "staff-next-actions@example.com", "password": "staff-password"})
    token = login_response.json()["access_token"]

    response = client.get("/v1/workspace/next-actions", headers={"Authorization": f"Bearer {token}"})

    assert response.status_code == 200
    actions = [item for bucket in response.json().values() for item in bucket]
    assert actions
    assert {item["action_type"] for item in actions} == {"upload_evidence"}


def test_workspace_machine_runs_refund_pipeline_for_owner(client: TestClient, db_session: Session) -> None:
    restaurant = Restaurant(name="Restaurant Machine", sender_email="claims@example.com")
    db_session.add(restaurant)
    db_session.flush()
    db_session.add(
        UberFinancialTransaction(
            restaurant_id=restaurant.id,
            uber_store_id="store-machine",
            uber_order_id="UBER-MACHINE-001",
            transaction_type="Commande non recue",
            amount=Decimal("-19.90"),
            currency="EUR",
            transaction_date=datetime(2026, 6, 1).date(),
            payout_reference="PAY-MACHINE-001",
            raw_payload_json={"description": "Commande non recue par le client"},
            imported_from="manager_export",
        )
    )
    db_session.commit()

    response = client.post(
        "/v1/workspace/machine/run",
        json={"trigger": "refunds", "sync_gmail": False, "run_autopilot": False},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["recipient_email"] == "restaurantsfrance@uber.com"
    stages = {stage["name"]: stage for stage in payload["stages"]}
    assert stages["deductions"]["created_count"] == 1
    assert stages["claim_orders"]["created_count"] == 1
    assert stages["gmail_sync"]["status"] == "skipped"
    dispute = db_session.scalar(select(UberCustomerRefundDispute))
    assert dispute is not None
    assert dispute.dispute_type == "order_not_received"
    assert dispute.claim_order_id is not None
    assert db_session.scalar(select(ClaimOrder).where(ClaimOrder.uber_order_number == "UBER-MACHINE-001")) is not None


def test_workspace_machine_creates_refund_case_from_single_stapled_ticket_proof(
    client: TestClient,
    db_session: Session,
) -> None:
    restaurant = Restaurant(name="Krousty Bat", sender_email="tiramisumaisonfrance@example.com")
    db_session.add(restaurant)
    db_session.commit()
    proof_text = (
        "Restaurant: Krousty Bat\n"
        "Client: Jean Test\n"
        "Commande: ABC123\n"
        "Date: 15/06/2026\n"
        "Montant total: 24,90 EUR\n"
        "Demande de remboursement - article manquant\n"
        "Ticket agrafe sur commande preparee et emballee\n"
    )
    preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("preuve-remboursement.jpg", proof_text.encode("utf-8"), "image/jpeg"))],
    ).json()
    confirm_response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [{"file_id": preview["files"][0]["id"], "action": "import_evidence_bulk"}],
        },
    )
    assert confirm_response.status_code == 200

    response = client.post(
        "/v1/workspace/machine/run",
        json={
            "trigger": "refunds",
            "smart_import_batch_id": preview["batch_preview_id"],
            "sync_gmail": False,
            "run_autopilot": False,
        },
    )

    assert response.status_code == 200
    stages = {stage["name"]: stage for stage in response.json()["stages"]}
    assert stages["proof_intake"]["created_count"] == 1
    order = db_session.scalar(select(ClaimOrder).where(ClaimOrder.uber_order_number == "ABC123"))
    assert order is not None
    assert order.restaurant_id == restaurant.id
    assert order.customer_name == "Jean Test"
    assert order.order_date.isoformat() == "2026-06-15"
    assert str(order.order_amount) == "24.90"
    assert order.status == "ready_to_send"
    dispute = db_session.scalar(select(UberCustomerRefundDispute).where(UberCustomerRefundDispute.claim_order_id == order.id))
    assert dispute is not None
    assert dispute.dispute_type == "missing_item"
    assert dispute.evidence_status == "complete"
    evidence = db_session.scalar(select(EvidenceFile).where(EvidenceFile.order_id == order.id))
    assert evidence is not None
    assert evidence.evidence_type == "receipt"


def test_workspace_machine_creates_refund_case_from_ocr_image_text(
    client: TestClient,
    db_session: Session,
    monkeypatch,
) -> None:
    restaurant = Restaurant(name="Krousty Bat", sender_email="tiramisumaisonfrance@example.com")
    db_session.add(restaurant)
    db_session.commit()
    monkeypatch.setattr(
        evidence_analysis_service,
        "extract_image_ocr_text",
        lambda _content: (
            "Restaurant: Krousty Bat\n"
            "Client: OCR Client\n"
            "Commande: OCR789\n"
            "Date: 15/06/2026\n"
            "Montant total: 18,70 EUR\n"
            "Remboursement client - commande non recue\n"
        ),
    )
    preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("photo-ticket.jpg", b"\xff\xd8\xff\xe0binary-image", "image/jpeg"))],
    ).json()
    confirm_response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [{"file_id": preview["files"][0]["id"], "action": "import_evidence_bulk"}],
        },
    )
    assert confirm_response.status_code == 200

    response = client.post(
        "/v1/workspace/machine/run",
        json={
            "trigger": "refunds",
            "smart_import_batch_id": preview["batch_preview_id"],
            "sync_gmail": False,
            "run_autopilot": False,
        },
    )

    assert response.status_code == 200
    order = db_session.scalar(select(ClaimOrder).where(ClaimOrder.uber_order_number == "OCR789"))
    assert order is not None
    assert order.customer_name == "OCR Client"
    assert str(order.order_amount) == "18.70"
    dispute = db_session.scalar(select(UberCustomerRefundDispute).where(UberCustomerRefundDispute.claim_order_id == order.id))
    assert dispute is not None
    assert dispute.dispute_type == "order_not_received"


def test_workspace_machine_reanalyzes_weak_current_batch_evidence_before_proof_intake(
    client: TestClient,
    db_session: Session,
    monkeypatch,
) -> None:
    restaurant = Restaurant(name="Frit Dodo", sender_email="frit@example.com")
    db_session.add(restaurant)
    db_session.commit()
    ocr_passes = iter(
        [
            "Ticket agrafe illisible\nRemboursement client\n",
            (
                "Restaurant: Frit Dodo\n"
                "Client: Client Reanalyse\n"
                "Commande: REOCR123\n"
                "Date: 15/06/2026\n"
                "Montant total: 21,80 EUR\n"
                "Demande de remboursement - commande non recue\n"
            ),
        ]
    )
    monkeypatch.setattr(
        evidence_analysis_service,
        "extract_image_ocr_text",
        lambda _content: next(ocr_passes),
    )
    preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("preuve-faible.jpg", b"\xff\xd8\xff\xe0binary-image", "image/jpeg"))],
    ).json()
    confirm_response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [{"file_id": preview["files"][0]["id"], "action": "import_evidence_bulk"}],
        },
    )
    assert confirm_response.status_code == 200
    assert db_session.scalar(select(ClaimOrder).where(ClaimOrder.uber_order_number == "REOCR123")) is None

    response = client.post(
        "/v1/workspace/machine/run",
        json={
            "trigger": "refunds",
            "smart_import_batch_id": preview["batch_preview_id"],
            "sync_gmail": False,
            "run_autopilot": False,
        },
    )

    assert response.status_code == 200
    stages = {stage["name"]: stage for stage in response.json()["stages"]}
    assert stages["evidence"]["processed_count"] >= 1
    assert stages["proof_intake"]["created_count"] == 1, " | ".join(stages["proof_intake"]["warnings"])
    order = db_session.scalar(select(ClaimOrder).where(ClaimOrder.uber_order_number == "REOCR123"))
    assert order is not None
    assert order.restaurant_id == restaurant.id
    assert order.customer_name == "Client Reanalyse"
    assert str(order.order_amount) == "21.80"


def test_workspace_machine_processes_targeted_smart_import_batch_even_when_not_recent(
    client: TestClient,
    db_session: Session,
) -> None:
    restaurant = Restaurant(name="Krousty Bat", sender_email="tiramisumaisonfrance@example.com")
    db_session.add(restaurant)
    db_session.commit()
    proof_text = (
        "Restaurant: Krousty Bat\n"
        "Client: Batch Cible\n"
        "Commande: TARGET123\n"
        "Date: 15/06/2026\n"
        "Montant total: 22,40 EUR\n"
        "Remboursement client - commande non recue\n"
    )
    preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("preuve-cible.jpg", proof_text.encode("utf-8"), "image/jpeg"))],
    ).json()
    confirm_response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [{"file_id": preview["files"][0]["id"], "action": "import_evidence_bulk"}],
        },
    )
    assert confirm_response.status_code == 200

    owner = db_session.scalar(select(User).where(User.email == "owner@example.com"))
    assert owner is not None
    for index in range(55):
        db_session.add(
            EvidenceImportBatch(
                uploaded_by_user_id=owner.id,
                restaurant_id=restaurant.id,
                original_filename=f"later-{index}.jpg",
                source_type="multi_file_upload",
                status="uploaded",
            )
        )
    db_session.commit()

    response = client.post(
        "/v1/workspace/machine/run",
        json={
            "trigger": "refunds",
            "smart_import_batch_id": preview["batch_preview_id"],
            "sync_gmail": False,
            "run_autopilot": False,
        },
    )

    assert response.status_code == 200
    stages = {stage["name"]: stage for stage in response.json()["stages"]}
    assert stages["evidence"]["processed_count"] >= 1
    assert stages["proof_intake"]["created_count"] == 1
    order = db_session.scalar(select(ClaimOrder).where(ClaimOrder.uber_order_number == "TARGET123"))
    assert order is not None
    assert order.customer_name == "Batch Cible"


def test_workspace_machine_resolves_proof_restaurant_from_existing_order_snapshot(
    client: TestClient,
    db_session: Session,
) -> None:
    restaurant = Restaurant(name="Asian Passion", sender_email="asian@example.com")
    db_session.add(restaurant)
    db_session.flush()
    snapshot = UberOrderSnapshot(
        restaurant_id=restaurant.id,
        uber_store_id="store-asian",
        uber_order_id="SNAP123",
        display_id="SNAP123",
        customer_name="Client Snapshot",
        current_state="completed",
        placed_at=utc_now(),
        order_total_amount=Decimal("21.40"),
        currency="EUR",
        raw_payload_json={"source": "test"},
        imported_from="manager_export",
    )
    db_session.add(snapshot)
    db_session.commit()

    preview = client.post(
        "/v1/smart-import/preview",
        files=[
            (
                "files",
                (
                    "preuve-remboursement.jpg",
                    b"Commande: SNAP123\nRemboursement client\nTicket agrafe",
                    "image/jpeg",
                ),
            )
        ],
    ).json()
    confirm_response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [{"file_id": preview["files"][0]["id"], "action": "import_evidence_bulk"}],
        },
    )
    assert confirm_response.status_code == 200

    response = client.post(
        "/v1/workspace/machine/run",
        json={
            "trigger": "manual",
            "smart_import_batch_id": preview["batch_preview_id"],
            "sync_gmail": False,
            "run_autopilot": False,
        },
    )

    assert response.status_code == 200
    order = db_session.scalar(select(ClaimOrder).where(ClaimOrder.uber_order_number == "SNAP123"))
    assert order is not None
    assert order.restaurant_id == restaurant.id
    assert order.customer_name == "Client Snapshot"
    assert str(order.order_amount) == "21.40"
    assert order.order_date == snapshot.placed_at.date()


def test_workspace_machine_hydrates_single_proof_from_historical_uber_import_row(
    client: TestClient,
    db_session: Session,
) -> None:
    restaurant = Restaurant(name="Frit Dodo", sender_email="frit@example.com")
    db_session.add(restaurant)
    db_session.flush()
    owner = db_session.scalar(select(User).where(User.email == "owner@example.com"))
    assert owner is not None
    batch = UberReportingImportBatch(
        uploaded_by_user_id=owner.id,
        original_filename="historique-uber.csv",
        report_type="combined_report",
        file_type="csv",
        status="parsed",
        total_rows=1,
        valid_rows=1,
    )
    db_session.add(batch)
    db_session.flush()
    db_session.add(
        UberReportingImportRow(
            batch_id=batch.id,
            row_number=1,
            raw_data={
                "Nom du restaurant": "Frit Dodo",
                "Id. de la commande": "HIST123",
                "Nom du client": "Client Historique",
                "Date de la commande": "15/06/2026",
                "Montant total": "28,60",
            },
            normalized_data={
                "restaurant_id": restaurant.id,
                "display_id": "HIST123",
                "customer_name": "Client Historique",
                "order_date": "2026-06-15",
                "order_amount": "28.60",
                "currency": "EUR",
            },
            status="valid",
        )
    )
    db_session.commit()

    proof_text = "Restaurant: Frit Dodo\nCommande: HIST123\nRemboursement client\nTicket agrafe\n"
    preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("preuve-historique.jpg", proof_text.encode("utf-8"), "image/jpeg"))],
    ).json()
    client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [{"file_id": preview["files"][0]["id"], "action": "import_evidence_bulk"}],
        },
    )

    response = client.post(
        "/v1/workspace/machine/run",
        json={
            "trigger": "refunds",
            "smart_import_batch_id": preview["batch_preview_id"],
            "sync_gmail": False,
            "run_autopilot": False,
        },
    )

    assert response.status_code == 200
    order = db_session.scalar(select(ClaimOrder).where(ClaimOrder.uber_order_number == "HIST123"))
    assert order is not None
    assert order.restaurant_id == restaurant.id
    assert order.customer_name == "Client Historique"
    assert order.order_date.isoformat() == "2026-06-15"
    assert str(order.order_amount) == "28.60"


def test_workspace_machine_extracts_excel_proof_cells(
    client: TestClient,
    db_session: Session,
) -> None:
    restaurant = Restaurant(name="Big Chicken Burger", sender_email="big@example.com")
    db_session.add(restaurant)
    db_session.commit()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Order accuracy"
    worksheet.append(["Restaurant", "Client", "Commande", "Montant total", "Motif"])
    worksheet.append(["Big Chicken Burger", "Excel Client", "XLX123", "17,90 EUR", "Remboursement client"])
    content = BytesIO()
    workbook.save(content)

    preview = client.post(
        "/v1/smart-import/preview",
        files=[
            (
                "files",
                (
                    "download.xlsx",
                    content.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            )
        ],
    ).json()
    confirm_response = client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [{"file_id": preview["files"][0]["id"], "action": "import_evidence_bulk"}],
        },
    )
    assert confirm_response.status_code == 200

    response = client.post(
        "/v1/workspace/machine/run",
        json={
            "trigger": "manual",
            "smart_import_batch_id": preview["batch_preview_id"],
            "sync_gmail": False,
            "run_autopilot": False,
        },
    )

    assert response.status_code == 200
    order = db_session.scalar(select(ClaimOrder).where(ClaimOrder.uber_order_number == "XLX123"))
    assert order is not None
    assert order.restaurant_id == restaurant.id
    assert order.customer_name == "Excel Client"
    assert str(order.order_amount) == "17.90"


def test_evidence_analysis_ignores_uber_report_definition_as_order_or_amount() -> None:
    text = (
        "Votre guide concernant les commandes incorrectes sur Uber Eats. "
        "Ce rapport a pour objectif de vous fournir des informations. "
        "Taux de commandes correctes total 100.00%."
    )

    assert evidence_analysis_service.extract_order_number(text) is None
    assert evidence_analysis_service.extract_amount(text) is None
    assert evidence_analysis_service.extract_amount("Montant total: 24,90 EUR") == Decimal("24.90")


def test_workspace_machine_creates_cancellation_case_from_single_stapled_ticket_proof(
    client: TestClient,
    db_session: Session,
) -> None:
    restaurant = Restaurant(name="Asian Passion", sender_email="asian@example.com")
    db_session.add(restaurant)
    db_session.commit()
    proof_text = (
        "Restaurant: Asian Passion\n"
        "Client: Client Annulation\n"
        "Commande: CAN123\n"
        "Date: 14/06/2026\n"
        "Montant total: 31.50 EUR\n"
        "Commande annulee apres preparation, ticket agrafe sur sac emballe\n"
    )
    preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("preuve-annulation.jpg", proof_text.encode("utf-8"), "image/jpeg"))],
    ).json()
    client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [{"file_id": preview["files"][0]["id"], "action": "import_evidence_bulk"}],
        },
    )

    response = client.post(
        "/v1/workspace/machine/run",
        json={
            "trigger": "cancellations",
            "smart_import_batch_id": preview["batch_preview_id"],
            "sync_gmail": False,
            "run_autopilot": False,
        },
    )

    assert response.status_code == 200
    stages = {stage["name"]: stage for stage in response.json()["stages"]}
    assert stages["proof_intake"]["created_count"] == 1
    order = db_session.scalar(select(ClaimOrder).where(ClaimOrder.uber_order_number == "CAN123"))
    assert order is not None
    assert order.restaurant_id == restaurant.id
    assert order.loss_type == "cancellation_not_compensated"
    assert order.customer_name == "Client Annulation"
    assert order.status == "ready_to_send"


def test_workspace_machine_keeps_incomplete_single_proof_unclassified(
    client: TestClient,
    db_session: Session,
) -> None:
    restaurant = Restaurant(name="Big Chicken Burger", sender_email="big@example.com")
    db_session.add(restaurant)
    db_session.commit()
    proof_text = (
        "Restaurant: Big Chicken Burger\n"
        "Client: Client Sans Montant\n"
        "Commande: MISS123\n"
        "Ticket agrafe mais montant illisible\n"
    )
    preview = client.post(
        "/v1/smart-import/preview",
        files=[("files", ("preuve-incomplete.jpg", proof_text.encode("utf-8"), "image/jpeg"))],
    ).json()
    client.post(
        "/v1/smart-import/confirm",
        json={
            "batch_preview_id": preview["batch_preview_id"],
            "files": [{"file_id": preview["files"][0]["id"], "action": "import_evidence_bulk"}],
        },
    )

    response = client.post(
        "/v1/workspace/machine/run",
        json={
            "trigger": "refunds",
            "smart_import_batch_id": preview["batch_preview_id"],
            "sync_gmail": False,
            "run_autopilot": False,
        },
    )

    assert response.status_code == 200
    stages = {stage["name"]: stage for stage in response.json()["stages"]}
    assert stages["proof_intake"]["created_count"] == 0
    assert stages["proof_intake"]["skipped_count"] >= 1
    assert "missing_montant" in " ".join(stages["proof_intake"]["warnings"])
    assert db_session.scalar(select(ClaimOrder).where(ClaimOrder.uber_order_number == "MISS123")) is None


def test_workspace_machine_repairs_historical_restaurant_misclassification(
    client: TestClient,
    db_session: Session,
) -> None:
    krousty = Restaurant(name="Krousty Bat", sender_email="krousty@example.com")
    asian = Restaurant(name="Asian Passion", sender_email="asian@example.com")
    db_session.add_all([krousty, asian])
    db_session.flush()
    db_session.add(
        UberStoreMapping(
            restaurant_id=asian.id,
            uber_store_id="store-asian-machine",
            uber_store_name="Asian Passion",
            active=True,
        )
    )
    snapshot = UberOrderSnapshot(
        restaurant_id=krousty.id,
        uber_store_id="store-asian-machine",
        uber_order_id="UBER-MACHINE-ASIAN-001",
        display_id="ASIAN-001",
        current_state="canceled",
        order_total_amount=Decimal("31.50"),
        currency="EUR",
        raw_payload_json={"uber_store_name": "Asian Passion", "uber_store_id": "store-asian-machine"},
        imported_from="manager_export",
    )
    transaction = UberFinancialTransaction(
        restaurant_id=krousty.id,
        uber_store_id="store-asian-machine",
        uber_order_id="UBER-MACHINE-ASIAN-001",
        transaction_type="customer_refund",
        amount=Decimal("-7.00"),
        currency="EUR",
        transaction_date=datetime(2026, 6, 1).date(),
        payout_reference="PAY-MACHINE-ASIAN-001",
        raw_payload_json={"uber_store_name": "Asian Passion", "uber_store_id": "store-asian-machine"},
        imported_from="manager_export",
    )
    order = ClaimOrder(
        restaurant_id=krousty.id,
        uber_order_number="UBER-MACHINE-ASIAN-001",
        order_amount=Decimal("31.50"),
        currency="EUR",
        status="missing_evidence",
    )
    db_session.add_all([snapshot, transaction, order])
    db_session.commit()

    response = client.post(
        "/v1/workspace/machine/run",
        json={
            "trigger": "manual",
            "sync_gmail": False,
            "run_autopilot": False,
            "run_historical_cleanup": True,
        },
    )

    assert response.status_code == 200
    stages = {stage["name"]: stage for stage in response.json()["stages"]}
    assert stages["historical_reclassification"]["created_count"] == 2
    db_session.refresh(snapshot)
    db_session.refresh(transaction)
    db_session.refresh(order)
    assert snapshot.restaurant_id == asian.id
    assert transaction.restaurant_id == asian.id
    assert order.restaurant_id == asian.id


def test_workspace_machine_reports_autopilot_disabled_without_failure(client: TestClient) -> None:
    response = client.post("/v1/workspace/machine/run", json={"sync_gmail": False, "run_autopilot": True})

    assert response.status_code == 200
    payload = response.json()
    stages = {stage["name"]: stage for stage in payload["stages"]}
    assert stages["autopilot"]["status"] == "skipped"
    assert "autopilot_disabled" in stages["autopilot"]["warnings"]


def test_workspace_machine_fast_go_skips_historical_cleanup_by_default(client: TestClient) -> None:
    response = client.post(
        "/v1/workspace/machine/run",
        json={"trigger": "manual", "sync_gmail": False, "run_autopilot": False},
    )

    assert response.status_code == 200
    stages = {stage["name"]: stage for stage in response.json()["stages"]}
    assert stages["historical_reclassification"]["status"] == "skipped"
    assert stages["historical_import_repair"]["status"] == "skipped"
    assert stages["historical_identity_hydration"]["status"] == "skipped"


def test_staff_cannot_run_workspace_machine(client: TestClient, db_session: Session) -> None:
    restaurant, _order = create_restaurant_order_and_task(db_session)
    staff = client.post(
        "/v1/users",
        json={
            "email": "staff-machine@example.com",
            "password": "staff-password",
            "full_name": "Staff Machine",
            "role": "staff",
            "active": True,
        },
    ).json()
    client.post(f"/v1/users/{staff['id']}/restaurants", json={"restaurant_id": restaurant.id})
    login_response = client.post("/v1/auth/login", json={"email": "staff-machine@example.com", "password": "staff-password"})
    token = login_response.json()["access_token"]

    response = client.post(
        "/v1/workspace/machine/run",
        json={"sync_gmail": False, "run_autopilot": False},
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 403


def test_evidence_analysis_rejects_numeric_customer_names() -> None:
    assert evidence_analysis_service.clean_customer_name("15.49") is None
    assert evidence_analysis_service.clean_customer_name("5.01 EUR") is None
    assert evidence_analysis_service.clean_customer_name("Client Test") == "Client Test"


def create_restaurant_order_and_task(db: Session) -> tuple[Restaurant, ClaimOrder]:
    owner = db.scalar(select(User).where(User.email == "owner@example.com"))
    restaurant = Restaurant(name="Restaurant Next Actions", sender_email="claims@example.com")
    db.add(restaurant)
    db.flush()
    order = ClaimOrder(
        restaurant_id=restaurant.id,
        uber_order_number="UBER-NEXT-001",
        customer_name="Client Test",
        order_amount=Decimal("125.00"),
        currency="EUR",
        notes="Annulation test",
        status="missing_evidence",
    )
    db.add(order)
    db.flush()
    db.add(
        EvidenceRequestTask(
            order_id=order.id,
            restaurant_id=restaurant.id,
            task_type="missing_receipt",
            required_evidence_type="receipt",
            status="pending",
            priority="urgent",
            title="Ticket a fournir",
            description="Ticket fictif requis",
            due_at=datetime.now(timezone.utc),
            reason="missing_receipt",
            created_by_user_id=owner.id if owner else None,
        )
    )
    db.commit()
    db.refresh(restaurant)
    db.refresh(order)
    return restaurant, order


def official_inaccurate_orders_csv() -> str:
    return (
        "Restaurant,Id. externe du restaurant,Pays,Code pays,Ville,Id. de la commande,UUID du processus,"
        "Heure de la commande,Heure d'acceptation par le marchand,Heure du remboursement,"
        "Problème avec la commande,Informations concernant le problème lié à l'article,Articles incorrects,"
        "Personnalisations incorrectes,Commentaires du client,Code de devise,Montant moyen des commandes,"
        "Client remboursé,Remboursement pris en charge par le commerçant,Remboursement non pris en charge par le commerçant,"
        "Type de commande honorée,Canal de commande,Marque Eats\n"
        "Krousty Bat,store-accuracy,France,FR,Paris,ORDER-ACCURACY-001,PROCESS-ACCURACY-001,"
        "2026-01-05 12:00,2026-01-05 12:02,2026-01-06 09:30,"
        "Article manquant,Burger Test,Burger Test,,Client Test,EUR,30.00,12.50,-12.50,0.00,"
        "Livraison,Uber Eats,Uber Eats\n"
    )
